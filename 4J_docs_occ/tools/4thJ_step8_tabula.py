#!/usr/bin/env python
"""
4J Step 8, work item 8.1 -- extract the TABULA archetype parameters for `es`,
`uk` and `it`, with a sheet-and-row reference on every value.

  usage: python 4thJ_step8_tabula.py <outputs_step8_dir>
         (expects the two workbooks in <outputs_step8_dir>/raw/)

WHAT THIS DOES AND DOES NOT DO
------------------------------
It reads TABULA and writes a parameter table. **It does not write an IDF.**
Geometry, zoning and construction layer build-up are decisions this file
deliberately does not take -- see THE HONESTY CLAUSE below, which is the part of
item 8.1 that actually protects the paper.

TWO WORKBOOKS, AND THE SPLIT BETWEEN THEM IS THE FIRST FINDING
--------------------------------------------------------------
  `tabula-values.xlsx`      4,028,656 B  md5 7347b2cae3c4d9f5ce78221e9d5fb832
      65 sheets. Constants, construction-year classes, construction assemblies,
      U-value classes, system efficiencies, and `Tab.BoundaryCond`.
      🔴 IT CONTAINS NO BUILDING ARCHETYPES. Searched every sheet for a first
      column matching a TABULA building-type code (`<CC>.N.<SFH|TH|MFH|AB>.…`):
      ZERO sheets carry one.
  `tabula-calculator.xlsx`  34,383,251 B md5 c99ddc9ffcb6dc0ae7391273d9619e37
      15 sheets. `Calc.Set.Building` is 3,287 data rows x 333 columns, one row
      per building VARIANT, and it is the only place the archetypes exist.

So the route recorded on 2026-08-20 -- "download `tabula-values.xlsx` … read
`Tab.Building.Constr`" -- reaches the CONSTRUCTIONS but not the BUILDINGS.
`Tab.Building.Constr` turns out to hold wall/roof/ceiling assemblies
(`ES.Wall.ReEx.01.01`), not archetypes. **Item 8.1 needs the 34 MB calculator,
which the 2026-08-20 entry explicitly recorded as NOT opened.** It is opened
here and its digest is pinned above.

VARIANTS -- AND WHY ONLY `.001` IS OURS
---------------------------------------
Each archetype appears as `<building>.001`, `.002`, `.003`: the existing state
and two refurbishment levels. 🔴 **Only variant `001` is the existing stock.**
Scoring a refurbished variant against a real diary would be comparing our
occupancy against a building that does not exist, so this extractor keeps `001`
and records how many rows it dropped.

🔴 `GB` IS GREAT BRITAIN, NOT THE UNITED KINGDOM
------------------------------------------------
TABULA's country code is `GB`; Northern Ireland is outside it. The fold is
labelled `uk` and its diaries are UK-wide. This is a declared limitation of the
same class as `D-S6-2`'s wave gap, and the emitted table carries `GB` in a
column of its own rather than being silently relabelled `uk`.

===========================================================================
THE HONESTY CLAUSE -- item 8.1's "record what TABULA does NOT give us"
===========================================================================
This is the part that matters, because an assumed value that is not written down
becomes a fact the moment someone reads the code. Checked against the workbooks,
not against a report:

**Searched every sheet header block of `tabula-values.xlsx` for:** `schedul`,
`hourly`, `sub-hour`, `tapping`, `window open`, `thermostat`, `set point`,
`setpoint`, `zoning`, `zone`, `occupan`, `appliance`, `plug`, `lighting`,
`draw`, `3d`. 🔴 **NOT ONE of these appears anywhere.** TABULA is a monthly
steady-state energy-balance method; it has no concept of a time series.

What it supplies INSTEAD, and this is sharper than "it does not supply
schedules", is a handful of ANNUAL CONSTANTS. But WHICH constants is itself a
finding, and it is the opposite of what `Tab.BoundaryCond` first suggests.

🔴 THE ARCHETYPES DO NOT USE THE NATIONAL BOUNDARY CONDITIONS. ALL THREE FOLDS
🔴 USE THE SAME `EU.SUH` / `EU.MUH` SET, AND THE DIFFERENCE IS LARGE.

`Tab.BoundaryCond` publishes national rows (`ES.SUH`, `ES.MUH`, `GB.Gen`,
`IT.SUH`, `IT.MUH`) AND an EU cross-country-comparison pair (`EU.SUH`,
`EU.MUH`). **Every one of the 102 archetype rows we keep points at the EU pair**
-- checked, not assumed: `Code_BoundaryCond` takes exactly two values across
`es`, `uk` and `it`, and it is `EU.SUH` for `SFH`/`TH` and `EU.MUH` for
`MFH`/`AB` in all three.

                        EU.SUH   EU.MUH  |  ES.SUH  ES.MUH  GB.Gen  IT.SUH  IT.MUH
    theta_i      degC      20       20   |    20      20      21      20      20
    F_red_htr1             0.9     0.95  |     1       1       1       1       1
    F_red_htr4             0.8     0.85  |     1       1       1       1       1
    n_air_use    1/h       0.4      0.4  |    0.4     0.4     0.59    0.3     0.3
    h_room       m         2.5      2.5  |    2.5     2.5     2.4     3.0     2.7
    phi_int      W/m2      3        3    |     3       3       4      2.8     4.1
    c_m       Wh/(m2K)     45       45   |    45      45     32.79    87      72
    q_w_nd   kWh/(m2 a)    10       15   |  11.09   21.76    15.8    14.5    17.3

So AS SHIPPED, and this is what the emitted tables carry:

  `phi_int` = **3.0 W/m2 in ALL THREE FOLDS**. 🔴 ONE NUMBER. No split into
      occupants / appliances / lighting, no time profile. **This is precisely the
      quantity our generated occupancy is meant to replace, so it is the
      injection point for the whole campaign**, and the uninjected control is the
      run that keeps it.
  `theta_i` = 20 degC everywhere; `n_air_use` = 0.4 /h everywhere -- window-opening
      behaviour is not modelled, it is folded into that one constant.
  `F_red_htr1/F_red_htr4` = 0.9/0.8 (SUH) and 0.95/0.85 (MUH). 🔴 SO THERE IS AN
      INTERMITTENT-HEATING REDUCTION AFTER ALL -- but it is a SCALAR applied to
      the transmission coefficient, not a setback schedule, and it is identical
      across our three folds. An EnergyPlus model that implements a real
      night-setback schedule is no longer computing TABULA's quantity.
  `q_w_nd` = 10 (SUH) / 15 (MUH) kWh/(m2 a). Annual totals, no tapping series.

🔴 WHY THIS MATTERS MORE THAN A UNITS QUIBBLE, AND WHY IT IS RECORDED HERE
RATHER THAN QUIETLY "IMPROVED" TO THE NATIONAL VALUES: on the EU set, every
non-geometric boundary condition is IDENTICAL across `es`, `uk` and `it`, so any
cross-country difference in a simulated result comes from geometry, U-values and
weather ALONE. On the national set it would additionally come from a 1 degC
set-point difference and a 0.3-vs-0.59 air-change difference, both
country-correlated -- i.e. confounded with exactly the LOCO signal this paper is
trying to measure. **Switching to the national rows would be a basis change and
is not taken here.**

🔴 A SECOND WRONG LABEL IN A PUBLISHED SOURCE. `Tab.BoundaryCond`'s unit row
gives `F_red_htr1` and `F_red_htr4` the unit **degC**. They are dimensionless
reduction factors -- the EU rows carry 0.9 / 0.8 and the German rows
0.8796296... / 0.7931034..., which are ratios and not temperatures. Same class as
`FINDING 47` and `FINDING 56` (`P139` in ISTAT's own tracciato): a published
label that contradicts its own column's values, catchable only by reading them.

⚪ Geometry: `Calc.Set.Building` gives ENVELOPE AREAS and a conditioned volume
(`A_Wall_1..3`, `A_Roof_1..2`, `A_Floor_1..2`, `A_Window_1..2`, the four
orientations plus horizontal, `V_C`, `n_Storey`) -- enough for a box model, and
NOT a 3D geometry or a zoning scheme. Those remain ours to assume and to declare.

⚪ `Remark_ConstructionYearClass` carries a descriptive label for **Spain only**
(`XIX century`, `Beginning of the century`, `Civil war`, `Improvement in the
Spanish economy`, `CTE-79`, `CTE 2006`). GB and IT carry none. `ES.05`'s label in
the file is `CTE-79`; `NBE-CT-79` is the historically correct name of the standard
and appears nowhere in the workbook.

⚪ **The licence was NOT verified here either.** Redistribution of derived tables
is still unchecked, and that check is owed before any of this is published.
"""

import csv
import io
import os
import sys

VALUES_XLSX = 'tabula-values.xlsx'
CALC_XLSX = 'tabula-calculator.xlsx'
VALUES_MD5 = '7347b2cae3c4d9f5ce78221e9d5fb832'
CALC_MD5 = 'c99ddc9ffcb6dc0ae7391273d9619e37'
VALUES_URL = 'https://episcope.eu/fileadmin/tabula/public/calc/tabula-values.xlsx'
CALC_URL = 'https://episcope.eu/fileadmin/tabula/public/calc/tabula-calculator.xlsx'

#: fold label -> TABULA country code. 🔴 `uk` maps to `GB`, which excludes
#: Northern Ireland; the mapping is written here so it cannot be assumed away.
FOLD_COUNTRY = {'es': 'ES', 'uk': 'GB', 'it': 'IT'}

#: The columns item 8.1 needs, by their TABULA header name. Anything not on this
#: list is not silently dropped -- the count of unused columns is reported.
WANTED = [
    'Code_BuildingVariant', 'Code_Building', 'Code_Country',
    'Code_BuildingType', 'Code_BuildingSizeClass', 'Code_ConstructionYearClass',
    'Code_ClimateRegion', 'Code_BoundaryCond', 'Code_StatusDataset',
    'Code_TypeVariant', 'Number_BuildingVariant',
    'A_C_Ref', 'V_C', 'n_Storey', 'n_Storey_effective', 'h_room',
    'A_Roof_1', 'A_Roof_2', 'A_Wall_1', 'A_Wall_2', 'A_Wall_3',
    'A_Floor_1', 'A_Floor_2', 'A_Window_1', 'A_Window_2', 'A_Door_1',
    'A_Window_Horizontal', 'A_Window_East', 'A_Window_South',
    'A_Window_West', 'A_Window_North',
    'U_Roof_1', 'U_Roof_2', 'U_Wall_1', 'U_Wall_2', 'U_Wall_3',
    'U_Floor_1', 'U_Floor_2', 'U_Window_1', 'U_Window_2', 'U_Door_1',
    'delta_U_ThermalBridging_Original',
    'phi_int', 'q_w_nd',
]

#: Read from `Tab.BoundaryCond`, keyed by `Code_BoundaryCond`.
BC_WANTED = ['theta_i', 'F_red_htr1', 'F_red_htr4', 'n_air_use', 'h_room',
             'phi_int', 'F_sh_hor', 'F_sh_vert', 'F_f', 'F_w', 'c_m', 'q_w_nd']


class TabulaError(RuntimeError):
    pass


def need(cond, msg):
    if not cond:
        raise TabulaError(msg)


def md5(path):
    import hashlib
    h = hashlib.md5()
    with open(path, 'rb') as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()


def load(path):
    try:
        import openpyxl
    except ImportError:
        raise TabulaError(
            "openpyxl is not installed. TABULA ships .xlsx and there is no CSV "
            "export; install it rather than re-typing values by hand.")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def build(outdir):
    raw = os.path.join(outdir, 'raw')
    log = []

    def say(s=''):
        log.append(s)
        print(s)

    vpath = os.path.join(raw, VALUES_XLSX)
    cpath = os.path.join(raw, CALC_XLSX)
    for p, want, url in ((vpath, VALUES_MD5, VALUES_URL), (cpath, CALC_MD5, CALC_URL)):
        need(os.path.exists(p),
             'missing %s\n  retrieve with:  curl -L -o "%s" "%s"' % (p, p, url))
        got = md5(p)
        need(got == want,
             '%s has md5 %s, expected %s. TABULA republishes without versioning, '
             'so a changed digest means the parameter table would silently stop '
             'matching its own provenance file. Refused.' % (p, got, want))
        say('%-24s md5 %s  OK' % (os.path.basename(p), got))

    # ---- construction-year classes, re-derived from the file -------------
    wv = load(vpath)
    rows = list(wv['Tab.ConstrYearClass'].iter_rows(max_row=400, values_only=True))
    hdr = list(rows[0])
    ix = dict((h, i) for i, h in enumerate(hdr) if h)
    bands = {}
    for r in rows[1:]:
        c = r[0]
        if not isinstance(c, str) or c[:2] not in ('ES', 'GB', 'IT') or c[2:3] != '.':
            continue
        bands[c] = (r[ix['ConstructionYearClass_FirstYear']],
                    r[ix['ConstructionYearClass_LastYear']],
                    r[ix['Remark_ConstructionYearClass']])
    say()
    say('--- construction-year classes, read from Tab.ConstrYearClass ---')
    for cc, n in (('ES', 6), ('GB', 8), ('IT', 8)):
        got = sorted(k for k in bands if k.startswith(cc))
        need(len(got) == n,
             '%s has %d construction-year classes, expected %d' % (cc, len(got), n))
        for k in got:
            lo, hi, rem = bands[k]
            say('  %-6s %5s-%-5s  %s' % (k, lo, hi, rem or ''))
    need(len(bands) == 22,
         'expected 22 ES/GB/IT construction-year classes, found %d' % len(bands))
    say('  22 of 22 bands present.')

    # ---- boundary conditions --------------------------------------------
    rows = list(wv['Tab.BoundaryCond'].iter_rows(max_row=200, values_only=True))
    hdr = list(rows[0])
    ix = dict((h, i) for i, h in enumerate(hdr) if h)
    bc, bc_eu = {}, {}
    for r in rows[1:]:
        c = r[0]
        if not isinstance(c, str) or c[2:3] != '.':
            continue
        if c[:2] in ('ES', 'GB', 'IT'):
            bc[c] = dict((k, r[ix[k]]) for k in BC_WANTED if k in ix)
        elif c[:2] == 'EU':
            bc_eu[c] = dict((k, r[ix[k]]) for k in BC_WANTED if k in ix)
    say()
    say('--- Tab.BoundaryCond, our three countries ---')
    say('  %-8s %s' % ('code', '  '.join('%s' % k for k in BC_WANTED)))
    for c in sorted(bc):
        say('  %-8s %s' % (c, '  '.join(str(bc[c].get(k)) for k in BC_WANTED)))
    say('  (national rows, for reference -- the archetypes do NOT use these)')
    say()
    say('--- Tab.BoundaryCond, the EU rows the archetypes ACTUALLY point at ---')
    say('  %-8s %s' % ('code', '  '.join('%s' % k for k in BC_WANTED)))
    for c in ('EU.SUH', 'EU.MUH'):
        need(c in bc_eu, '%s is absent from Tab.BoundaryCond' % c)
        say('  %-8s %s' % (c, '  '.join(str(bc_eu[c].get(k)) for k in BC_WANTED)))
    # The honesty clause above quotes these sixteen numbers. If TABULA republishes
    # with different ones the clause silently becomes fiction, so they are pinned.
    EXPECT = {'EU.SUH': dict(theta_i=20, F_red_htr1=0.9, F_red_htr4=0.8,
                             n_air_use=0.4, h_room=2.5, phi_int=3, c_m=45, q_w_nd=10),
              'EU.MUH': dict(theta_i=20, F_red_htr1=0.95, F_red_htr4=0.85,
                             n_air_use=0.4, h_room=2.5, phi_int=3, c_m=45, q_w_nd=15)}
    for c, want in sorted(EXPECT.items()):
        for k, v in sorted(want.items()):
            got = bc_eu[c].get(k)
            need(abs(float(got) - v) < 1e-9,
                 "%s.%s is %r, but this module's honesty clause states %r. A "
                 "republished workbook must not silently change the numbers the "
                 "provenance file quotes." % (c, k, got, v))
    say('  all 16 quoted EU boundary-condition values match the file.')
    say('  phi_int = 3.0 W/m2 in ALL THREE FOLDS -- the injection point.')
    say('  F_red_htr is 0.9/0.8 and 0.95/0.85, NOT 1: an intermittent-heating')
    say('  SCALAR exists, and it is identical across the three folds.')

    # ---- the archetypes, from the calculator ----------------------------
    wc = load(cpath)
    need('Calc.Set.Building' in wc.sheetnames,
         'Calc.Set.Building is absent from %s' % CALC_XLSX)
    it = wc['Calc.Set.Building'].iter_rows(values_only=True)
    hdr = list(next(it))
    for _ in range(4):
        next(it)
    ix = dict((h, i) for i, h in enumerate(hdr) if isinstance(h, str))
    missing = [k for k in WANTED if k not in ix]
    need(not missing, 'Calc.Set.Building is missing columns %s' % missing)

    kept = dict((f, []) for f in FOLD_COUNTRY)
    n_seen = n_variant = n_noclass = 0
    noclass = []
    for r in it:
        code = r[0]
        if not isinstance(code, str) or len(code) < 4 or code[2] != '.':
            continue
        cc = code[:2]
        fold = next((f for f, c in FOLD_COUNTRY.items() if c == cc), None)
        if fold is None:
            continue
        n_seen += 1
        if not code.endswith('.001'):          # refurbishment variants
            n_variant += 1
            continue
        # 🔴 Four `ES.TestRegion.MUH*.SyAv` rows carry a real `A_C_Ref`
        # (1034.6 - 1499.6 m2) and NO construction-year class. They are TABULA's
        # own test rows and they sit under `Code_StatusDataset = Typology`, so
        # neither the status column nor a non-null area check excludes them. Any
        # extraction that keys on the country code alone ships them, and Spain
        # then reports SEVEN construction-year classes where the census axis has
        # SIX. Dropped on the construction-year class, and counted.
        cyc = r[ix['Code_ConstructionYearClass']]
        if not (isinstance(cyc, str) and cyc[:2] == cc and cyc[2:3] == '.'):
            n_noclass += 1
            noclass.append(code)
            continue
        kept[fold].append(dict((k, r[ix[k]]) for k in WANTED))

    say()
    say('--- Calc.Set.Building ---')
    say('  ES/GB/IT rows seen           %d' % n_seen)
    say('  refurbishment variants dropped %d  (kept only `.001`, the existing state)'
        % n_variant)
    say('  columns used %d of %d' % (len(WANTED), len(hdr)))
    say('  rows with NO construction-year class dropped %d  %s' % (n_noclass, noclass))
    need(n_noclass == 4 and all(x.startswith('ES.TestRegion.') for x in noclass),
         'expected exactly the four ES.TestRegion rows to lack a construction-year '
         'class, got %d: %s. A new unclassified row is a new contaminant and this '
         'refuses rather than dropping it quietly.' % (n_noclass, noclass))
    for fold in ('es', 'uk', 'it'):
        need(kept[fold],
             'fold %r kept zero archetypes; a parameter table over an empty set '
             'is not a parameter table' % fold)
        periods = sorted(set(str(x['Code_ConstructionYearClass']) for x in kept[fold]))
        types = sorted(set(str(x['Code_BuildingType']) for x in kept[fold]))
        say('  %-3s (%s)  %3d archetypes  types %s  periods %d'
            % (fold, FOLD_COUNTRY[fold], len(kept[fold]), types, len(periods)))
        used_bc = sorted(set(str(x['Code_BoundaryCond']) for x in kept[fold]))
        need(used_bc == ['EU.MUH', 'EU.SUH'],
             '%s archetypes point at boundary conditions %s, not the EU pair this '
             'module documents. The honesty clause would then be describing a '
             'parameter set the tables do not use.' % (fold, used_bc))
        expect = {'es': 6, 'uk': 8, 'it': 8}[fold]
        need(len(periods) == expect,
             '%s covers %d construction-year classes, but Tab.ConstrYearClass '
             'declares %d' % (fold, len(periods), expect))
        import collections as _c
        grid = _c.Counter((str(x['Code_ConstructionYearClass']),
                           str(x['Code_BuildingType']).split('.')[2]
                           if str(x['Code_BuildingType']).count('.') >= 2 else '?')
                          for x in kept[fold])
        shapes = sorted(set(k[1] for k in grid))
        say('      building types %s ; %d of %d type x period cells populated'
            % (shapes, len(grid), len(shapes) * expect))

    # ---- write ----------------------------------------------------------
    os.makedirs(outdir, exist_ok=True)
    for fold in ('es', 'uk', 'it'):
        p = os.path.join(outdir, 'archetype_parameters_%s.csv' % fold)
        with io.open(p, 'w', newline='', encoding='utf-8') as fh:
            w = csv.DictWriter(fh, fieldnames=WANTED)
            w.writeheader()
            for row in sorted(kept[fold], key=lambda x: x['Code_BuildingVariant']):
                w.writerow(row)
            fh.write('# fold %s -> TABULA country %s%s\n'
                     % (fold, FOLD_COUNTRY[fold],
                        '  (GREAT BRITAIN: Northern Ireland excluded)'
                        if fold == 'uk' else ''))
            fh.write('# source Calc.Set.Building in %s md5 %s from %s\n'
                     % (CALC_XLSX, CALC_MD5, CALC_URL))
            fh.write('# variant .001 only (existing state); %d refurbishment '
                     'variants dropped across all three folds\n' % n_variant)
        say('WROTE %s  (%d archetypes)' % (p, len(kept[fold])))
    return log


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print('usage: python 4thJ_step8_tabula.py <outputs_step8_dir>', file=sys.stderr)
        sys.exit(2)
    try:
        build(sys.argv[1])
    except TabulaError as exc:
        print('EXTRACTION REFUSED: %s' % exc, file=sys.stderr)
        sys.exit(1)
