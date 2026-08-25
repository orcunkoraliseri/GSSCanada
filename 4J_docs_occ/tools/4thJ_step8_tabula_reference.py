# -*- coding: utf-8 -*-
"""4J Step 8 item 8.3 --- TABULA's OWN annual heating energy need, per archetype.

WHY THIS FILE EXISTS
--------------------
`G8.7` reads "per-archetype EUI vs published band, as-modelled = PASS, empirical
= INFO", and on 2026-08-20 `FINDING 44` recorded that `G8.1`-`G8.4` name no
reference series at all.  That finding examined three candidates --- measured
data, the flat-4.0 foil, the uninjected control --- and rejected all three.

A fourth was never examined, and it is sitting in the workbook this project
already downloaded: **`tabula-calculator.xlsx` publishes `q_h_nd`, the annual
energy need for heating, in kWh/(m2.a), for every building variant** --- which
includes all 86 distinct archetype codes work item 8.1 built.  Measured here,
not assumed: 86 of 86 matched.

That is an AS-MODELLED reference of the SAME quantity: TABULA's seasonal method
against our hourly one, on the same envelope, the same boundary conditions and
the same climate region.  It is exactly the comparison `G8.7` was written for,
and it is NOT the inversion trap of `FINDING 44`(b)/(c) --- it does not become
easier to satisfy when the paper's claim is true, because it is computed from
the archetype and never from our schedules.

WHAT IS EXTRACTED, AND WHAT IS DELIBERATELY EXTRACTED WITH IT
-------------------------------------------------------------
`q_h_nd` alone would be a number without a basis.  The components come with it,
because the control campaign has to be able to SAY WHY it differs rather than
just report that it does:

  q_ht_tr, q_ht_ve   transmission and ventilation heat transfer
  q_sol, q_int       solar and internal gains
  eta_h_gn           gain utilisation factor
  F_red_temp         TABULA's intermittent-heating reduction (a SCALAR, see
                     FINDING 57 --- we deliberately did not implement it as a
                     schedule, so it is a KNOWN difference, not a discrepancy)
  Theta_e, theta_i   the climate and set-point TABULA itself used
  A_C_Ref            the reference area every kWh/m2 in this project divides by

and one DERIVED column that is the whole story of `FINDING 119` again:

  heating_hours_implied = q_int * 1000 / phi_int

TABULA counts gains only inside its heating season, so this recovers the length
of that season in hours.  For `ES.ME` it comes back at ~538 h --- 22.4 days,
which is `Tab.AuxCalc.Climate`'s own `HeatingDays` for that region.  An hourly
model with a 20 C set-point heats whenever the zone falls below it, all year.
**The two methods are not measuring the same season**, and any comparison that
does not say so out loud is quoting an artefact.

VARIANT
-------
`Code_BuildingVariant` ends in `.001` / `.002` / `.003` = existing condition /
usual refurbishment / advanced refurbishment.  Work item 8.1 built the EXISTING
condition (`Number_BuildingVariant == 1`, verified against the parameter tables'
own variant column), so variant 1 is the row taken and the tool ABORTS rather
than guessing if a code's variant 1 is missing.

Output: `Step8_docs/outputs_step8/tabula_reference.csv`.
"""
import csv
import io
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
PROJ = os.path.dirname(HERE)
BASE = os.path.join(PROJ, "Step8_docs", "outputs_step8")
XLSX = os.path.join(BASE, "raw", "tabula-calculator.xlsx")
IDF_MANIFEST = os.path.join(BASE, "archetype_idf_manifest.csv")
OUT = os.path.join(BASE, "tabula_reference.csv")

SHEET = "Calc.Set.Building"

# Column names, taken from the sheet's own header row -- never by position.
WANT = ["Code_BuildingVariant", "Code_Building", "Code_Country", "Code_ClimateRegion",
        "Code_BoundaryCond", "Code_StatusDataset", "Number_BuildingVariant",
        "A_C_Ref", "Theta_e", "theta_i", "F_red_temp",
        "q_ht_tr", "q_ht_ve", "q_ht", "q_sol", "q_int", "eta_h_gn", "q_h_nd", "q_w_nd"]

PHI_INT = 3.0  # W/m2, EU.SUH / EU.MUH, identical in all three folds (FINDING 57)


def load_codes():
    rows = list(csv.DictReader(io.open(IDF_MANIFEST, encoding="utf-8")))
    if not rows:
        sys.exit("archetype_idf_manifest.csv is empty")
    return rows


def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required (python 3.13 interpreter has it)")

    if not os.path.exists(XLSX):
        sys.exit("missing %s" % XLSX)

    man = load_codes()
    codes = sorted(set(r["code"] for r in man))
    fold_of = {}
    for r in man:
        fold_of.setdefault(r["code"], r["fold"])

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit("sheet %r not in %s" % (SHEET, XLSX))
    ws = wb[SHEET]

    it = ws.iter_rows(min_row=1, values_only=True)
    hdr = list(next(it))
    idx = {}
    for w in WANT:
        if w not in hdr:
            sys.exit("column %r is not in %s -- refusing to guess a position" % (w, SHEET))
        idx[w] = hdr.index(w)

    by_code = {}
    for r in it:
        cv = r[idx["Code_BuildingVariant"]]
        if not isinstance(cv, str) or cv.count(".") < 6:
            continue          # header block, template row, blank
        rec = {}
        for w in WANT:
            rec[w] = r[idx[w]]
        by_code.setdefault(rec["Code_Building"], []).append(rec)

    missing = [c for c in codes if c not in by_code]
    if missing:
        sys.exit("%d archetype codes are not in %s: %s"
                 % (len(missing), SHEET, ", ".join(missing[:5])))

    out = []
    for c in codes:
        v1 = [x for x in by_code[c] if str(x["Number_BuildingVariant"]) in ("1", "1.0")]
        if len(v1) != 1:
            sys.exit("code %s has %d rows at Number_BuildingVariant == 1 -- "
                     "the existing-condition variant must be unique" % (c, len(v1)))
        rec = dict(v1[0])
        rec["fold"] = fold_of[c]
        qi = rec["q_int"]
        rec["heating_hours_implied"] = (float(qi) * 1000.0 / PHI_INT) if qi else ""
        rec["heating_days_implied"] = ((float(qi) * 1000.0 / PHI_INT) / 24.0) if qi else ""
        rec["n_variants_published"] = len(by_code[c])
        out.append(rec)

    cols = (["fold", "Code_Building", "Code_BuildingVariant", "Number_BuildingVariant",
             "n_variants_published"]
            + [w for w in WANT if w not in ("Code_Building", "Code_BuildingVariant",
                                            "Number_BuildingVariant")]
            + ["heating_hours_implied", "heating_days_implied"])

    with io.open(OUT, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for rec in sorted(out, key=lambda x: (x["fold"], x["Code_Building"])):
            w.writerow(rec)

    print("codes in the IDF manifest      : %d distinct (%d IDFs)" % (len(codes), len(man)))
    print("matched in %-20s: %d" % (SHEET, len(out)))
    print("written                        : %s" % os.path.relpath(OUT, PROJ))
    for f in ("es", "uk", "it"):
        sel = [x for x in out if x["fold"] == f]
        if not sel:
            continue
        q = sorted(float(x["q_h_nd"]) for x in sel if x["q_h_nd"] is not None)
        hd = [float(x["heating_days_implied"]) for x in sel if x["heating_days_implied"] != ""]
        print("  %-3s n=%2d  q_h_nd kWh/(m2.a) min %8.2f  median %8.2f  max %8.2f | "
              "TABULA heating days %.1f"
              % (f, len(sel), q[0], q[len(q) // 2], q[-1],
                 sum(hd) / len(hd) if hd else float("nan")))


if __name__ == "__main__":
    main()
