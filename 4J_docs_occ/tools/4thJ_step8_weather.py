#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Work item 8.2 --- the weather files, acquired and SELECTED BY MEASUREMENT.

WHAT THIS IMPLEMENTS, AND WHAT IT REPLACES
------------------------------------------
Section 6 item 6 was ruled on 2026-08-21 as *diary-survey-year actual weather*:
`es` 2009-2010, `uk` 2014-2015, `it` 2013-2014.  That ruling was reversed by the
author on 2026-08-25 --- a typical year built from many years is acceptable, and
the exact year is not required.  This file implements the reversal.  See
`Step8_docs/docs/2026-08-25_D-S8-4_weather-basis-and-station-selection.md`.

The reversal is not a loss.  The 2026-08-21 entry recorded the price of the
actual-year ruling in its own words: under it "two things differ at once" across
folds --- the country AND the meteorological year --- so a cross-fold difference
in demand could no longer be attributed to the LOCO transfer.  A single shared
TMYx base period removes that confound.  The pre-registered reporting rule that
came with the old ruling ("any cross-fold comparison of absolute demand must name
the meteorological year") is therefore satisfied trivially: all three folds name
the same base period, 2009-2023.

THE BASE PERIOD IS THE SAME FOR ALL THREE FOLDS, ON PURPOSE
-----------------------------------------------------------
climate.onebuilding.org publishes TMYx in several vintages (..., 2004-2018,
2007-2021, 2009-2023, 2011-2025).  Mixing vintages across folds would reintroduce
exactly the asymmetry the reversal removes.  **TMYx.2009-2023** is used for all
three, and it is not an arbitrary pick: it is the only published 15-year window
that CONTAINS all three original fieldwork windows (es 2009-2010, uk 2014-2015,
it 2013-2014).  So the abandoned actual years are inside the base period of the
typical year that replaced them.

THE STATION IS MEASURED, NOT QUOTED
------------------------------------
TABULA does not give coordinates.  It gives climate-region CODES --- `ES.ME`,
`GB.Temperate`, `IT.MidClim` --- and, in `Tab.AuxCalc.Climate` of
`tabula-calculator.xlsx`, the twelve monthly mean external temperatures
`theta_e_01..12` those codes stand for.  Twelve published numbers per region is
enough to SELECT a station rather than assume one:

    score(station) = RMSE over the 12 months of ( EPW monthly mean dry bulb
                                                  - TABULA theta_e_MM )

and the station with the lowest score wins.  Annual global horizontal irradiation
is measured too and reported, but is NOT part of the score --- see below.

TWO THINGS THAT ARE MEASURED AND DELIBERATELY NOT SCORED
---------------------------------------------------------
* `HeatingDays` and `Theta_e_HeatingSeason`.  TABULA publishes both, and they
  look comparable, and they are NOT.  TABULA derives them from MONTHLY means by a
  fractional formula (`ES.ME` gets `HeatingDays_01 = 21.5`, not 31), while the
  obvious EPW statistic counts DAILY means below the base temperature.  For
  `ES.ME` the two differ by a factor of three on the same climate --- 22 against
  72 --- purely because daily scatter crosses a 12 C threshold that a monthly
  mean of 11.0 C never crosses.  Scoring on it would have rejected every real
  Spanish coastal station.  Recorded so that nobody re-derives the same false
  gap: `FINDING 119`.
* Annual global horizontal irradiation.  All 45 candidates come out ABOVE
  TABULA's published `I_Sol_Year_Hor`, by +2 % to +20 %, so it separates stations
  much less than it looks and it carries a systematic offset of its own.  It is
  recorded per station in the selection report and per fold in the manifest, and
  the residue is stated in the decision brief rather than optimised away.

OUTPUTS
-------
  outputs_step8/weather/<three .epw files>
  outputs_step8/weather/_cache/<the .zip as downloaded, kept for the md5>
  outputs_step8/weather_manifest.csv
  outputs_step8/weather_selection_report.json

Usage:
    python tools/4thJ_step8_weather.py Step8_docs/outputs_step8
    python tools/4thJ_step8_weather.py Step8_docs/outputs_step8 --offline
"""

import argparse
import csv
import hashlib
import io
import json
import os
import sys
import urllib.request
import zipfile

BASE_PERIOD = "TMYx.2009-2023"
SUFFIX = "_" + BASE_PERIOD + ".zip"
ROOT = "https://climate.onebuilding.org/WMO_Region_6_Europe/%s/"
COUNTRY_DIR = {"es": "ESP_Spain", "uk": "GBR_United_Kingdom", "it": "ITA_Italy"}

# The archetype tables carry `GB.Temperate`; `tabula-calculator.xlsx` carries the
# same region as `GB.England-Temperate`.  One region, two spellings, in two
# workbooks of the same release --- `FINDING 118`.  Aliased here explicitly so
# the lookup cannot silently miss and fall back to a default.
REGION_ALIAS = {"GB.Temperate": "GB.England-Temperate"}

# Candidate shortlists.  Not every station in the country: the ones whose region
# is the one TABULA names, plus a deliberate CONTROL per fold that should lose
# (Madrid for a Mediterranean region, Rome for the Italian Zone E) so the score
# can be seen discriminating rather than merely agreeing with a prior.
CANDIDATES = {
    "es": [
        "CT_Catalonia/ESP_CT_Barcelona-El.Prat.AP.081810",
        "CT_Catalonia/ESP_CT_Barcelona.081800",
        "VC_Valencia/ESP_VC_Valencia.AP.082840",
        "VC_Valencia/ESP_VC_Valencia.Viveros.082850",
        "VC_Valencia/ESP_VC_Castellon.Almazora.082860",
        "MC_Murcia/ESP_MC_Alicante.AP.083600",
        "MC_Murcia/ESP_MC_Murcia.084300",
        "AN_Andalusia/ESP_AN_Malaga.AP.084820",
        "AN_Andalusia/ESP_AN_Almeria.AP.084870",
        "IB_Balearic_Islands/ESP_IB_Palma.083010",
        "MD_Madrid/ESP_MD_Madrid-Barajas-Suarez.AP.082210",   # control, inland
    ],
    "uk": [
        "ENG_England/GBR_ENG_London-Heathrow.Intl.AP.037720",
        "ENG_England/GBR_ENG_London-Gatwick.AP.037760",
        "ENG_England/GBR_ENG_London.Wea.Ctr-St.James.Park.037700",
        "ENG_England/GBR_ENG_London-Stansted.AP.036830",
        "ENG_England/GBR_ENG_Birmingham.AP.035340",
        "ENG_England/GBR_ENG_Manchester.AP.033340",
        "ENG_England/GBR_ENG_Nottingham.Watnall.033540",
        "ENG_England/GBR_ENG_Leeds.Bradford.AP.033463",
        "ENG_England/GBR_ENG_Bristol.AP.037243",
        "ENG_England/GBR_ENG_RAF.Waddington.033770",
    ],
    "it": [
        "LM_Lombardy/ITA_LM_Milano-Linate.AP.160800",
        "LM_Lombardy/ITA_LM_Milano-Malpensa.AP.160660",
        "LM_Lombardy/ITA_LM_Milano-Bergamo.Intl.AP.160760",
        "LM_Lombardy/ITA_LM_Ghedi.AB.160880",
        "LM_Lombardy/ITA_LM_Brescia-Montichiari.AP.162593",
        "PM_Piedmont/ITA_PM_Torino-Caselle.AP.160590",
        "PM_Piedmont/ITA_PM_Torino.Venaria.160600",
        "PM_Piedmont/ITA_PM_Casale.Monferrato.160680",
        "PM_Piedmont/ITA_PM_Cuneo-Levaldigi.AP.161170",
        "VD_Valle_d-Aosta/ITA_VD_Novara-Cameri.AB.160640",
        "ER_Emilia-Romagna/ITA_ER_Bologna-Marconi.AP.161400",
        "ER_Emilia-Romagna/ITA_ER_Parma-Verdi.AP.161300",
        "ER_Emilia-Romagna/ITA_ER_Piacenza-San.Damiano.AB.160840",
        "ER_Emilia-Romagna/ITA_ER_Ferrara.AP.161380",
        "VN_Veneto/ITA_VN_Verona.Catullo.AP.160900",
        "VN_Veneto/ITA_VN_Venezia-Polo.AP.161050",
        "VN_Veneto/ITA_VN_Padova.AP.160950",
        "VN_Veneto/ITA_VN_Treviso.AP.160990",
        "VN_Veneto/ITA_VN_Istrana.AB.160980",
        "FV_Friuli-Venezia_Giulia/ITA_FV_Udine-Rivolto.AB.160450",
        "FV_Friuli-Venezia_Giulia/ITA_FV_Aviano.AFB.160360",
        "TC_Tuscany/ITA_TC_Firenze-Peretola.AP.161700",
        "LZ_Lazio/ITA_LZ_Roma-Ciampino.AP.162390",            # control, Zone D
    ],
}

# EPW column indices actually read here.  EnergyPlus reads more; the selftest
# checks the ones a run cannot proceed without.
C_MONTH, C_DAY, C_DRYBULB, C_GHI = 1, 2, 6, 13


# ------------------------------------------------------------------ TABULA
def region_codes(base):
    """The climate region each fold's archetypes actually carry, read from the
    parameter tables rather than assumed."""
    out = {}
    for fold in ("es", "uk", "it"):
        p = os.path.join(base, "archetype_parameters_%s.csv" % fold)
        seen = set()
        for r in csv.DictReader(io.open(p, encoding="utf-8-sig")):
            v = r.get("Code_ClimateRegion")
            if v:
                seen.add(v)
        if len(seen) != 1:
            sys.exit("fold %s carries %d climate regions (%s); one EPW per fold "
                     "is only correct while that is 1" % (fold, len(seen), seen))
        out[fold] = seen.pop()
    return out


def tabula_targets(base, codes):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is not installed; tabula-calculator.xlsx cannot be read.")
    wb = openpyxl.load_workbook(
        os.path.join(base, "raw", "tabula-calculator.xlsx"),
        read_only=True, data_only=True)
    ws = wb["Tab.AuxCalc.Climate"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                             max_col=ws.max_column, values_only=True))
    hdr = [("" if c is None else str(c)) for c in rows[0]]

    def col(name):
        for i, h in enumerate(hdr):
            if h == name:
                return i
        sys.exit("TABULA climate sheet has no column %r" % name)

    out = {}
    for fold, code in codes.items():
        want = REGION_ALIAS.get(code, code)
        hit = [r for r in rows if r[0] == want]
        if len(hit) != 1:
            sys.exit("climate region %r (from %r) matches %d rows in "
                     "Tab.AuxCalc.Climate" % (want, code, len(hit)))
        r = hit[0]
        out[fold] = {
            "code_in_archetypes": code,
            "code_in_calculator": want,
            "name": r[col("Name_ClimateRegion")],
            "theta_month": [float(r[col("theta_e_%02d" % m)]) for m in range(1, 13)],
            "ghi_year_kwh": float(r[col("I_Sol_Year_Hor")]),
            "heating_days_tabula": float(r[col("HeatingDays")]),
            "theta_heating_tabula": float(r[col("Theta_e_HeatingSeason")]),
        }
    return out


# ------------------------------------------------------------------- EPW
def fetch(url, dest, offline):
    if os.path.exists(dest) and os.path.getsize(dest) > 10000:
        return False
    if offline:
        sys.exit("--offline but %s is not cached" % os.path.basename(dest))
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=180) as r:
        data = r.read()
    io.open(dest, "wb").write(data)
    return True


def read_epw(path):
    txt = io.open(path, encoding="latin-1").read().split("\n")
    loc = txt[0].strip()
    rows = [ln.strip().split(",") for ln in txt[8:]
            if ln.strip() and len(ln.strip().split(",")) >= 16]
    return loc, rows


def measure(rows):
    """Monthly mean dry bulb, annual GHI, and the two TABULA-incomparable
    heating-season statistics (kept for the record, never scored)."""
    day, ghi = {}, 0.0
    for f in rows:
        t = float(f[C_DRYBULB])
        if t >= 99.0:
            raise ValueError("missing dry bulb sentinel in EPW")
        day.setdefault((int(f[C_MONTH]), int(f[C_DAY])), []).append(t)
        ghi += float(f[C_GHI])
    dmean = {k: sum(v) / len(v) for k, v in day.items()}
    heat = [v for v in dmean.values() if v <= 12.0]
    return {
        "n_rows": len(rows),
        "n_days": len(dmean),
        "hours_per_day": sorted(set(len(v) for v in day.values())),
        "theta_month": [sum(v for (m, _), v in dmean.items() if m == mm)
                        / sum(1 for (m, _) in dmean if m == mm)
                        for mm in range(1, 13)],
        "theta_year": sum(dmean.values()) / len(dmean),
        "ghi_year_kwh": ghi / 1000.0,
        "heating_days_daily_basis": len(heat),
        "theta_heating_daily_basis": (sum(heat) / len(heat)) if heat else None,
    }


def rmse(a, b):
    return (sum((x - y) ** 2 for x, y in zip(a, b)) / float(len(a))) ** 0.5


def md5(path):
    h = hashlib.md5()
    with io.open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("base")
    ap.add_argument("--offline", action="store_true",
                    help="use the cached zips only; fail rather than download")
    a = ap.parse_args()

    base = os.path.abspath(a.base)
    wdir = os.path.join(base, "weather")
    cdir = os.path.join(wdir, "_cache")
    for d in (wdir, cdir):
        if not os.path.isdir(d):
            os.makedirs(d)

    codes = region_codes(base)
    targets = tabula_targets(base, codes)
    print("climate regions: %s" % codes)

    report = {"base_period": BASE_PERIOD,
              "source": "climate.onebuilding.org",
              "score": "RMSE of 12 monthly mean dry-bulb vs TABULA theta_e_01..12",
              "targets": targets, "candidates": {}}
    manifest = []

    for fold in ("es", "uk", "it"):
        t = targets[fold]
        cands = []
        for rel in CANDIDATES[fold]:
            url = ROOT % COUNTRY_DIR[fold] + rel + SUFFIX
            zp = os.path.join(cdir, os.path.basename(rel) + SUFFIX)
            fetch(url, zp, a.offline)
            with zipfile.ZipFile(zp) as z:
                names = [n for n in z.namelist() if n.lower().endswith(".epw")]
                if len(names) != 1:
                    sys.exit("%s holds %d EPW files" % (zp, len(names)))
                epw_name = os.path.basename(names[0])
                epw_path = os.path.join(cdir, epw_name)
                io.open(epw_path, "wb").write(z.read(names[0]))
            loc, rows = read_epw(epw_path)
            m = measure(rows)
            m.update({
                "rel": rel, "url": url, "epw": epw_name,
                "zip_md5": md5(zp), "epw_md5": md5(epw_path),
                "location_header": loc,
                "rmse_theta_month": rmse(m["theta_month"], t["theta_month"]),
                "d_ghi_pct": 100.0 * (m["ghi_year_kwh"] / t["ghi_year_kwh"] - 1.0),
            })
            cands.append(m)
        cands.sort(key=lambda c: c["rmse_theta_month"])
        report["candidates"][fold] = cands
        win = cands[0]

        # install the winner
        src = os.path.join(cdir, win["epw"])
        dst = os.path.join(wdir, win["epw"])
        io.open(dst, "wb").write(io.open(src, "rb").read())

        loc = win["location_header"].split(",")
        manifest.append({
            "fold": fold,
            "climate_region": codes[fold],
            "climate_region_calculator": t["code_in_calculator"],
            "climate_region_name": t["name"],
            "epw": win["epw"],
            "base_period": BASE_PERIOD,
            "station": loc[1] if len(loc) > 1 else "",
            "wmo": loc[5] if len(loc) > 5 else "",
            "lat": loc[6] if len(loc) > 6 else "",
            "lon": loc[7] if len(loc) > 7 else "",
            "tz": loc[8] if len(loc) > 8 else "",
            "elev_m": loc[9] if len(loc) > 9 else "",
            "source_url": win["url"],
            "zip_md5": win["zip_md5"],
            "epw_md5": md5(dst),
            "rmse_theta_month": "%.4f" % win["rmse_theta_month"],
            "runner_up": cands[1]["epw"],
            "runner_up_rmse": "%.4f" % cands[1]["rmse_theta_month"],
            "n_candidates": len(cands),
            "theta_year_c": "%.3f" % win["theta_year"],
            "ghi_year_kwh": "%.1f" % win["ghi_year_kwh"],
            "ghi_vs_tabula_pct": "%+.2f" % win["d_ghi_pct"],
            "heating_days_daily_basis": win["heating_days_daily_basis"],
            "tabula_heating_days_note": "not comparable, FINDING 119",
        })
        print("%s -> %-46s rmse %.3f   runner-up %.3f   GHI %+.1f%%"
              % (fold, win["epw"], win["rmse_theta_month"],
                 cands[1]["rmse_theta_month"], win["d_ghi_pct"]))

    mp = os.path.join(base, "weather_manifest.csv")
    with io.open(mp, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(manifest[0].keys()))
        w.writeheader()
        for r in manifest:
            w.writerow(r)
    rp = os.path.join(base, "weather_selection_report.json")
    io.open(rp, "w", encoding="utf-8").write(json.dumps(report, indent=1))
    print("\nwrote %s\nwrote %s" % (mp, rp))


if __name__ == "__main__":
    main()
