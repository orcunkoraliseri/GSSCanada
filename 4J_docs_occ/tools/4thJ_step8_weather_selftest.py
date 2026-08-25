#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Selftest for work item 8.2 --- the weather files.

Two halves, and as in 8.1 the second is the one that matters.

  W. THE FILES AND THE SELECTION.  Re-derives, from the EPW bytes and from
     `tabula-calculator.xlsx`, everything `weather_manifest.csv` claims.  It
     never trusts a stored column: `A5` in the 8.1 selftest was a no-op for
     exactly that reason, and the lesson is applied here from the start.

  B. ENERGYPLUS PROVENANCE.  Runs one archetype per fold against that fold's
     EPW and reads the `Site:Location` line E+ wrote back out of `eplusout.eio`.
     A manifest row saying "this run used Birmingham" is only a claim; the WMO
     number E+ echoes is the measurement.  `3J`'s inherited `PLATFORM` field is
     the precedent --- a provenance field can only be tested by changing the
     thing it claims to record.

Usage:
    python tools/4thJ_step8_weather_selftest.py Step8_docs/outputs_step8
    python tools/4thJ_step8_weather_selftest.py Step8_docs/outputs_step8 --no-eplus
"""

import argparse
import csv
import hashlib
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile

TOL_RMSE = 1e-4          # K, on re-deriving the recorded score
TOL_DEG = 0.02           # degrees, on lat/lon round-tripped through E+
TOL_ELEV = 0.6           # m, E+ prints elevation to 2 dp from the EPW

# EPW missing-value sentinels for the fields a run cannot proceed without.
# (index, sentinel-at-or-above, name)
SENTINELS = [
    (6, 99.0, "dry bulb"),
    (7, 99.0, "dew point"),
    (8, 999.0, "relative humidity"),
    (9, 999999.0, "atmospheric pressure"),
    (13, 9999.0, "global horizontal radiation"),
    (14, 9999.0, "direct normal radiation"),
    (15, 9999.0, "diffuse horizontal radiation"),
    (20, 999.0, "wind direction"),
    (21, 999.0, "wind speed"),
]

OK, BAD = [], []


def check(name, cond, detail=""):
    (OK if cond else BAD).append((name, detail))
    print("  %-62s %s%s" % (name, "ok" if cond else "FAILED",
                            ("   " + detail) if detail and not cond else ""))


def md5(path):
    h = hashlib.md5()
    with io.open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def read_epw(path):
    txt = io.open(path, encoding="latin-1").read().split("\n")
    rows = [ln.strip().split(",") for ln in txt[8:]
            if ln.strip() and len(ln.strip().split(",")) >= 22]
    return txt[0].strip(), rows


def monthly_theta(rows):
    day = {}
    for f in rows:
        day.setdefault((int(f[1]), int(f[2])), []).append(float(f[6]))
    dm = {k: sum(v) / len(v) for k, v in day.items()}
    out = []
    for mm in range(1, 13):
        vals = [v for (m, _), v in dm.items() if m == mm]
        out.append(sum(vals) / len(vals))
    return out, dm


def tabula_theta(base, code_calc):
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(base, "raw", "tabula-calculator.xlsx"),
                                read_only=True, data_only=True)
    ws = wb["Tab.AuxCalc.Climate"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                             max_col=ws.max_column, values_only=True))
    hdr = [("" if c is None else str(c)) for c in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    hit = [r for r in rows if r[0] == code_calc]
    if len(hit) != 1:
        return None
    return [float(hit[0][idx["theta_e_%02d" % m]]) for m in range(1, 13)]


def rmse(a, b):
    return (sum((x - y) ** 2 for x, y in zip(a, b)) / float(len(a))) ** 0.5


# --------------------------------------------------------------- half W
def half_w(base):
    mp = os.path.join(base, "weather_manifest.csv")
    rp = os.path.join(base, "weather_selection_report.json")
    rows = list(csv.DictReader(io.open(mp, encoding="utf-8")))
    rep = json.load(io.open(rp, encoding="utf-8"))
    wdir = os.path.join(base, "weather")

    print("\nW. THE FILES AND THE SELECTION  (%d folds)" % len(rows))

    check("W1  one weather file per fold, exactly es / uk / it",
          sorted(r["fold"] for r in rows) == ["es", "it", "uk"],
          str([r["fold"] for r in rows]))

    missing, mismatch = [], []
    for r in rows:
        p = os.path.join(wdir, r["epw"])
        if not os.path.exists(p):
            missing.append(r["epw"])
        elif md5(p) != r["epw_md5"]:
            mismatch.append(r["epw"])
    check("W2  every declared EPW is on disk and its md5 matches the manifest",
          not missing and not mismatch,
          "missing %s; md5 %s" % (missing, mismatch))
    if missing:
        return

    shape, sentinel, order = [], [], []
    theta = {}
    for r in rows:
        loc, data = read_epw(os.path.join(wdir, r["epw"]))
        if len(data) != 8760:
            shape.append("%s has %d rows" % (r["fold"], len(data)))
            continue
        m, dm = monthly_theta(data)
        theta[r["fold"]] = (m, loc)
        hours = {}
        for f in data:
            hours[(int(f[1]), int(f[2]))] = hours.get((int(f[1]), int(f[2])), 0) + 1
        if len(dm) != 365 or set(hours.values()) != {24}:
            shape.append("%s has %d days, hours/day %s"
                         % (r["fold"], len(dm), sorted(set(hours.values()))))
        months = [int(f[1]) for f in data]
        if months != sorted(months):
            order.append(r["fold"])
        for i, lim, nm in SENTINELS:
            n = sum(1 for f in data if float(f[i]) >= lim)
            if n:
                sentinel.append("%s: %d missing %s" % (r["fold"], n, nm))

    check("W3  every EPW is 8760 rows, 365 days, exactly 24 hours per day",
          not shape, "; ".join(shape))
    check("W4  no missing-value sentinel in any field EnergyPlus reads",
          not sentinel, "; ".join(sentinel[:4]))
    check("W5  hours run Jan to Dec in order, no month out of sequence",
          not order, str(order))

    periods = set(r["base_period"] for r in rows)
    infile = set()
    for r in rows:
        tok = [t for t in r["epw"].replace(".epw", "").split("_") if t.startswith("TMYx")]
        infile.add(tok[0] if tok else "?")
    check("W6  all three folds share ONE base period, and the filenames say so",
          len(periods) == 1 and infile == periods,
          "manifest %s vs filenames %s" % (periods, infile))

    hdr_bad = []
    for r in rows:
        loc = theta[r["fold"]][1].split(",")
        if loc[5].strip() != r["wmo"].strip():
            hdr_bad.append("%s wmo %s vs %s" % (r["fold"], loc[5], r["wmo"]))
        if abs(float(loc[6]) - float(r["lat"])) > 1e-6 or \
           abs(float(loc[7]) - float(r["lon"])) > 1e-6:
            hdr_bad.append("%s coords" % r["fold"])
    check("W7  the manifest's station, WMO and coordinates are the EPW's own header",
          not hdr_bad, "; ".join(hdr_bad))

    score_bad, argmin_bad, degenerate = [], [], []
    for r in rows:
        tgt = tabula_theta(base, r["climate_region_calculator"])
        if tgt is None:
            score_bad.append("%s: no TABULA row for %s"
                             % (r["fold"], r["climate_region_calculator"]))
            continue
        got = rmse(theta[r["fold"]][0], tgt)
        if abs(got - float(r["rmse_theta_month"])) > TOL_RMSE:
            score_bad.append("%s: re-derived %.4f vs recorded %s"
                             % (r["fold"], got, r["rmse_theta_month"]))
        cands = rep["candidates"][r["fold"]]
        best = min(cands, key=lambda c: c["rmse_theta_month"])
        if best["epw"] != r["epw"]:
            argmin_bad.append("%s: installed %s, argmin is %s"
                              % (r["fold"], r["epw"], best["epw"]))
        worst = max(c["rmse_theta_month"] for c in cands)
        if worst < 2.0 * best["rmse_theta_month"]:
            degenerate.append("%s: worst %.2f vs best %.2f"
                              % (r["fold"], worst, best["rmse_theta_month"]))

    check("W8  the recorded score is re-derivable from the EPW and TABULA itself",
          not score_bad, "; ".join(score_bad))
    check("W9  the installed station IS the argmin of its fold's candidate set",
          not argmin_bad, "; ".join(argmin_bad))
    check("W10 the score discriminates --- worst candidate is >= 2x the best",
          not degenerate, "; ".join(degenerate))

    # Not pass/fail: how safe the winner's margin is.  Recorded because the UK
    # margin is thin enough that the choice is nearly a tie (`FINDING 120`).
    print("\n   W11  NOT pass/fail --- the winning margin, measured:")
    for r in rows:
        print("        %s  %-46s %.3f   runner-up %-40s %s"
              % (r["fold"], r["epw"][:46], float(r["rmse_theta_month"]),
                 r["runner_up"][:40], r["runner_up_rmse"]))


# --------------------------------------------------------------- half B
def half_b(base, eplus):
    mp = os.path.join(base, "weather_manifest.csv")
    rows = list(csv.DictReader(io.open(mp, encoding="utf-8")))
    arch = os.path.join(base, "archetypes")
    idfs = sorted(os.listdir(arch))
    print("\nB. ENERGYPLUS PROVENANCE  (%d runs)" % len(rows))

    tmp = tempfile.mkdtemp(prefix="4j_s82_")
    fail, prov = [], []
    try:
        for r in rows:
            pick = [f for f in idfs if f.startswith(r["fold"] + "_")]
            if not pick:
                fail.append("%s: no archetype IDF" % r["fold"])
                continue
            d = os.path.join(tmp, r["fold"])
            os.makedirs(d)
            shutil.copy(os.path.join(arch, pick[0]), os.path.join(d, "in.idf"))
            epw = os.path.join(base, "weather", r["epw"])
            p = subprocess.run([eplus, "-w", epw, "-d", d, "-r",
                                os.path.join(d, "in.idf")],
                               cwd=d, stdout=subprocess.PIPE,
                               stderr=subprocess.STDOUT)
            err = os.path.join(d, "eplusout.err")
            sev = 0
            if os.path.exists(err):
                sev = io.open(err, encoding="utf-8",
                              errors="replace").read().count("** Severe")
            if p.returncode != 0 or sev:
                fail.append("%s rc=%s sev=%s" % (r["fold"], p.returncode, sev))
                continue
            eio = os.path.join(d, "eplusout.eio")
            line = ""
            for ln in io.open(eio, encoding="utf-8", errors="replace"):
                if ln.startswith("Site:Location,"):
                    line = ln.strip()
                    break
            if not line:
                prov.append("%s: E+ wrote no Site:Location" % r["fold"])
                continue
            f = line.split(",")
            name, lat, lon, _tz, elev = f[1], float(f[2]), float(f[3]), f[4], float(f[5])
            if ("WMO#=" + r["wmo"].strip()) not in name:
                prov.append("%s: E+ ran WMO %r, manifest says %s"
                            % (r["fold"], name, r["wmo"]))
            if abs(lat - float(r["lat"])) > TOL_DEG or \
               abs(lon - float(r["lon"])) > TOL_DEG:
                prov.append("%s: E+ ran %.3f/%.3f, manifest %s/%s"
                            % (r["fold"], lat, lon, r["lat"], r["lon"]))
            if abs(elev - float(r["elev_m"])) > TOL_ELEV:
                prov.append("%s: E+ elevation %.2f vs manifest %s"
                            % (r["fold"], elev, r["elev_m"]))
            print("   %s  %s" % (r["fold"], name))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    check("B1  one archetype per fold runs on its own EPW, 0 severe errors",
          not fail, "; ".join(fail))
    check("B2  the station EnergyPlus echoes back IS the manifest's station",
          not prov, "; ".join(prov))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("base")
    ap.add_argument("--eplus", default=r"C:\EnergyPlusV24-2-0\energyplus.exe")
    ap.add_argument("--no-eplus", action="store_true")
    a = ap.parse_args()
    base = os.path.abspath(a.base)

    half_w(base)
    if not a.no_eplus:
        if os.path.exists(a.eplus):
            half_b(base, a.eplus)
        else:
            print("\nB. SKIPPED --- EnergyPlus not found at %s" % a.eplus)

    print("\n%d ok, %d FAILED" % (len(OK), len(BAD)))
    for n, d in BAD:
        print("   FAILED  %s  %s" % (n, d))
    sys.exit(1 if BAD else 0)


if __name__ == "__main__":
    main()
