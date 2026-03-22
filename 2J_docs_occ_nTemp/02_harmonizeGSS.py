"""
02_harmonizeGSS.py

Step 2 of the Occupancy Modeling Pipeline: Data Harmonization.
Transforms the four raw Step 1 CSV outputs into a unified, cross-cycle-compatible schema.
"""

import pathlib
import pandas as pd
import openpyxl

INPUT_DIR = "outputs_step1/"
OUTPUT_DIR = "outputs_step2/"

CYCLES = [2005, 2010, 2015, 2022]

SENTINEL_MAP = {
    "KOL": {7, 8, 9, 98, 99},
    "TOTINC": {98, 99},
    "HRSWRK": {96, 97, 98, 99},
    "NOCS": {96, 97, 98, 99},
    # LFTAG sentinels are handled inside recode_lftag per-cycle before this map runs.
    "MARSTH": {8, 9, 99},
    # "MODE": {99},
    "WKSWRK": {96, 97, 98, 99},
    # COW sentinels are handled inside recode_cow() per-cycle (values differ across cycles).
}

if __name__ == "__main__":
    pass

# --- PHASE C: DEMOGRAPHIC RECODES ---


def recode_sex(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Rename only processing; ensure values are {1, 2}."""
    return df


def recode_marsth(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Unified values: {1, 2, 3, 4, 5, 6} + NaN."""
    if cycle in (2005, 2010):
        df["MARSTH"] = df["MARSTH"].replace({8: pd.NA, 9: pd.NA})
    elif cycle == 2022:
        df["MARSTH"] = df["MARSTH"].replace({99: pd.NA})
    return df


def recode_agegrp(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """1-7 unified scheme. No changes needed."""
    return df


def recode_lftag(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Collapse labour-force status to a 3-category scheme comparable across all cycles.
    Target scheme:
        1 = Working at paid job (any hours/type; includes vacation, maternity)
        2 = Going to school (student, no paid employment)
        3 = Not employed (retired, HH work, looking for work, illness, other)

    2005/2010 LFSGSS:
        1 (FT employed)       → 1
        2 (PT employed)       → 1
        3 (Student+employed)  → 1
        4 (Student only)      → 2
        5 (No employment)     → 3  (undifferentiated; covers retired/HH work/other)
        8/9 (NS/DK)           → NaN

    2015 ACT7DAYS:
        1 (Working)           → 1
        2 (Looking for work)  → 3
        3 (School)            → 2
        4 (HH work/caring)    → 3
        5 (Retired)           → 3
        6 (Other)             → 3
        97/98/99 (NS/RF/DK)   → NaN

    2022 ACT7DAYC:
        1 (Working)           → 1
        2 (School)            → 2
        3 (HH work/caring)    → 3
        4 (Retired)           → 3
        5 (Other)             → 3
        9 (NS)                → NaN
    """
    if cycle in (2005, 2010):
        df["LFTAG"] = df["LFTAG"].replace({8: pd.NA, 9: pd.NA})
        df["LFTAG"] = df["LFTAG"].replace({1: 1, 2: 1, 3: 1, 4: 2, 5: 3})
    elif cycle == 2015:
        df["LFTAG"] = df["LFTAG"].replace({97: pd.NA, 98: pd.NA, 99: pd.NA})
        df["LFTAG"] = df["LFTAG"].replace({1: 1, 2: 3, 3: 2, 4: 3, 5: 3, 6: 3})
    elif cycle == 2022:
        df["LFTAG"] = df["LFTAG"].replace({9: pd.NA})
        df["LFTAG"] = df["LFTAG"].replace({1: 1, 2: 2, 3: 3, 4: 3, 5: 3})
    return df


def recode_pr(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """No remap needed. 2005 stays as REGION (1-5), others as PRV (10-59)."""
    return df


def recode_cma(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """LUC_RST 1-3. No remap needed."""
    return df


def recode_hhsize(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Collapse 2005-2015 code 6 into 5."""
    if cycle in (2005, 2010, 2015):
        df["HHSIZE"] = df["HHSIZE"].replace({6: 5})
    return df


def recode_hrswrk(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Map continuous hours to discrete brackets matching 2022 (8 bins)."""
    df["HRSWRK_RAW"] = df["HRSWRK"]
    if cycle in (2005, 2010, 2015):
        # Null out sentinel/skip codes BEFORE binning.
        # Valid range is 0–75 for all three cycles; sentinel codes (97/99.6/99.7/99.8/99.9)
        # are all > 75 and must not be binned into category 8 (60+ hours).
        df["HRSWRK"] = pd.to_numeric(df["HRSWRK"], errors="coerce")
        df["HRSWRK"] = df["HRSWRK"].where(df["HRSWRK"] <= 75)
        # 1: Under 30, 2: 30-34, 3: 35-39, 4: 40-44, 5: 45-49, 6: 50-54, 7: 55-59, 8: 60+
        bins = [-1, 29.99, 34.99, 39.99, 44.99, 49.99, 54.99, 59.99, 200]
        labels = [1, 2, 3, 4, 5, 6, 7, 8]
        mask = df["HRSWRK"].notna()
        df.loc[mask, "HRSWRK"] = pd.cut(
            df.loc[mask, "HRSWRK"], bins=bins, labels=labels, right=True
        ).astype("float64")
    elif cycle == 2022:
        # 2022 already in 1-8 bins. Just ensure it is float.
        df["HRSWRK"] = pd.to_numeric(df["HRSWRK"], errors="coerce")
    return df


def recode_kol(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Collapse language categories to 3 buckets: 1=English, 2=French, 3=Other/Both/Multiple."""
    if cycle in (2005, 2010):
        # LANCH: 1=Eng, 2=Fre, 3=Other, 4/5/6/7=Multiple
        mapping = {1: 1, 2: 2, 3: 3, 4: 3, 5: 3, 6: 3, 7: 3}
        df["KOL"] = df["KOL"].map(mapping)
    elif cycle in (2015, 2022):
        # LAN_01: 1=Eng, 2=Fre, 3=Both, 4=Neither
        mapping = {1: 1, 2: 2, 3: 3, 4: 3}
        df["KOL"] = df["KOL"].map(mapping)
    return df


def derive_mode(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Hierarchical priority mapping."""
    if cycle == 2005:
        df["MODE"] = pd.NA
    elif cycle == 2010:
        # Hierarchical derived from CTW checkboxes
        def get_priority_2010(row):
            def is_checked(val):
                return val in (1, 1.0, "1", "1.0", "Yes", "yes", "YES")

            if is_checked(row.get("CTW_Q140_C01")):
                return 1
            if is_checked(row.get("CTW_Q140_C02")):
                return 2
            if is_checked(row.get("CTW_Q140_C04")) or is_checked(
                row.get("CTW_Q140_C05")
            ):
                return 3
            if is_checked(row.get("CTW_Q140_C08")):
                return 4
            if is_checked(row.get("CTW_Q140_C07")):
                return 5
            if (
                is_checked(row.get("CTW_Q140_C06"))
                or is_checked(row.get("CTW_Q140_C09"))
                or is_checked(row.get("CTW_Q140_C03"))
            ):
                return 6
            return pd.NA

        df["MODE"] = df.apply(get_priority_2010, axis=1)
    elif cycle == 2015:
        # Hierarchical priority
        def get_priority_2015(row):
            def is_checked(val):
                return val in (1, 1.0, "1", "1.0", "Yes", "yes", "YES")

            if is_checked(row.get("CTW_140A")):
                return 1
            if is_checked(row.get("CTW_140B")):
                return 2
            if is_checked(row.get("CTW_140C")) or is_checked(row.get("CTW_140D")):
                return 3
            if is_checked(row.get("CTW_140G")):
                return 4
            if is_checked(row.get("CTW_140F")):
                return 5
            if (
                is_checked(row.get("CTW_140E"))
                or is_checked(row.get("CTW_140H"))
                or is_checked(row.get("CTW_140I"))
            ):
                return 6
            return pd.NA

        df["MODE"] = df.apply(get_priority_2015, axis=1)
    elif cycle == 2022:
        df["MODE"] = df["MODE"].replace({99: pd.NA})
    return df


def recode_totinc(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Income harmonized to buckets. Discretize CRA values in 2022."""
    df["TOTINC_RAW"] = df["TOTINC"]
    if cycle in (2005, 2010):
        df["TOTINC_SOURCE"] = "SELF"
        mapping = {
            1: 1,
            2: 1,
            3: 1,
            4: 1,
            5: 1,
            6: 2,
            7: 2,
            8: 3,
            9: 3,
            10: 4,
            11: 5,
            12: 5,
        }
        df["TOTINC"] = df["TOTINC"].map(mapping)
    elif cycle == 2015:
        df["TOTINC_SOURCE"] = "SELF"
        mapping = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 5, 7: 5}
        df["TOTINC"] = df["TOTINC"].map(mapping)
    elif cycle == 2022:
        df["TOTINC_SOURCE"] = "CRA"
    return df


def recode_cow(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Harmonize Class of Worker to 3-category scheme:
        1 = Employee (paid worker for someone else)
        2 = Self-employed (with or without employees) or unpaid family worker
        NaN = Not applicable / not at paid work / sentinel
    """
    if "COW" not in df.columns:
        return df
    if cycle in (2005, 2010):
        # MAR_Q172: 1=employee, 2=self-employed, 3=unpaid family, 7=not asked, 8/9=sentinel
        df["COW"] = df["COW"].replace({1: 1, 2: 2, 3: 2, 7: pd.NA, 8: pd.NA, 9: pd.NA})
    elif cycle == 2015:
        # WHW_110: 1=self-employed (w/ paid help), 2=employee, 6=not applicable, 7/8/9=sentinel
        df["COW"] = df["COW"].replace({1: 2, 2: 1, 6: pd.NA, 7: pd.NA, 8: pd.NA, 9: pd.NA})
    elif cycle == 2022:
        # WET_120: 1=employee, 2=self-emp w/ employees, 3=self-emp w/o employees, 6=N/A, 9=sentinel
        df["COW"] = df["COW"].replace({1: 1, 2: 2, 3: 2, 6: pd.NA, 9: pd.NA})
    return df


# --- PHASE D: ACTIVITY CODE CROSSWALK ---

ACTIVITY_EXCEL = (
    "references_activityCodes/Data Harmonization_activityCategories - execution.xlsx"
)
ACT_SHEET_MAP = {
    2005: "2005codebook",
    2010: "2010codebook",
    2015: "2015codebook",
    2022: "2022codebook",
}

ACT_LABELS = {
    1: "Work & Related",
    2: "Household Work & Maintenance",
    3: "Caregiving & Help",
    4: "Purchasing Goods & Services",
    5: "Sleep & Naps & Resting",
    6: "Eating & Drinking",
    7: "Personal Care",
    8: "Education",
    9: "Socializing",
    10: "Passive Leisure",
    11: "Active Leisure",
    12: "Community & Volunteer",
    13: "Travel",
    14: "Miscellaneous / Idle",
}


def build_activity_crosswalks(excel_path: str) -> dict[int, dict]:
    """Build {cycle_year: {raw_code: category_num}} from the Excel."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    crosswalks = {}
    for cycle_year, sheet_name in ACT_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None or type(row[0]) == str:
                continue
            category = int(row[0])
            raw_str = str(row[1]).replace(",", ".")
            raw_code = float(raw_str) if cycle_year == 2010 else int(float(raw_str))
            lookup[raw_code] = category
            if cycle_year == 2010:
                lookup[int(raw_code)] = category
        crosswalks[cycle_year] = lookup
    return crosswalks


def apply_activity_crosswalk(
    df: pd.DataFrame, cycle: int, crosswalk: dict
) -> pd.DataFrame:
    """Map raw activity codes to 14-category scheme."""
    raw_col = "ACTCODE" if cycle in (2005, 2010) else "TUI_01"
    if raw_col not in df.columns:
        return df

    df["occACT_raw"] = df[raw_col]

    if cycle == 2010:
        df["occACT"] = (
            pd.to_numeric(df[raw_col], errors="coerce").astype(float).map(crosswalk)
        )
        mask = df["occACT"].isna() & df[raw_col].notna()
        if mask.any():
            fallback_map = (
                pd.to_numeric(df.loc[mask, raw_col], errors="coerce")
                .fillna(-1)
                .astype(int)
                .map(crosswalk)
            )
            df.loc[mask, "occACT"] = fallback_map
    else:
        df["occACT"] = pd.to_numeric(df[raw_col], errors="coerce").map(crosswalk)

    if cycle in (2005, 2010):
        df.loc[df[raw_col] == 995, "occACT"] = 10
    elif cycle == 2015:
        df.loc[df[raw_col] == 95, "occACT"] = 14
    elif cycle == 2022:
        df.loc[df[raw_col] == 9999, "occACT"] = 14

    df["occACT_label"] = df["occACT"].map(ACT_LABELS)
    return df


def validate_activity_crosswalk(
    df: pd.DataFrame, cycle: int, raw_col: str
) -> None:
    """Log unmapped raw activity codes with their frequencies."""
    nan_mask = df["occACT"].isna() & df[raw_col].notna()
    if not nan_mask.any():
        print(f"  [{cycle}] All activity codes mapped. ✅")
        return

    unmapped = df.loc[nan_mask, raw_col].value_counts()
    n_eps = nan_mask.sum()
    n_resp = df.loc[nan_mask, "occID"].nunique()
    print(f"  [{cycle}] ⚠️  {n_eps} unmapped episodes in {n_resp} respondents:")
    for code, count in unmapped.items():
        pct = count / len(df) * 100
        print(f"    {raw_col}={code}: {count} episodes ({pct:.2f}%)")


# --- PHASE E: PRESENCE & CO-PRESENCE ---

PRESENCE_EXCEL = (
    "references_Pre_coPre_Codes/Data Harmonization_presenceCategories - execution.xlsx"
)
PRES_SHEET_MAP = {
    2005: "2005-2010codebook",
    2010: "2005-2010codebook",
    2015: "2015codebook",
    2022: "2022codebook",
}


def build_presence_crosswalks(excel_path: str) -> dict[int, dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    crosswalks = {}
    for cycle_year, sheet_name in PRES_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None or type(row[0]) == str:
                continue
            unified = int(row[0])
            raw_code = int(float(row[1]))
            lookup[raw_code] = unified
        crosswalks[cycle_year] = lookup
    return crosswalks


def apply_presence_crosswalk(
    df: pd.DataFrame, cycle: int, crosswalk: dict
) -> pd.DataFrame:
    raw_col = "PLACE" if cycle in (2005, 2010) else "LOCATION"
    if raw_col not in df.columns:
        return df
    df["occPRE_raw"] = df[raw_col]
    df["occPRE"] = pd.to_numeric(df[raw_col], errors="coerce").map(crosswalk)
    df["AT_HOME"] = (df["occPRE"] == 1).astype(int)
    return df


COPRESENCE_MAP = {
    2005: {
        "ALONE": "Alone",
        "SPOUSE": "Spouse",
        "CHILDHSD": "Children",
        "FRIENDS": "friends",
        "OTHFAM": "otherHHs",
        "OTHERS": "others",
        "PARHSD": "parents",
        "MEMBHSD": "otherInFAMs",
    },
    2010: {
        "ALONE": "Alone",
        "SPOUSE": "Spouse",
        "CHILDHSD": "Children",
        "FRIENDS": "friends",
        "OTHFAM": "otherHHs",
        "OTHERS": "others",
        "PARHSD": "parents",
        "MEMBHSD": "otherInFAMs",
    },
    2015: {
        "TUI_06A": "Alone",
        "TUI_06B": "Spouse",
        "TUI_06C": "Children",
        "TUI_06H": "friends",
        "TUI_06G": "otherHHs",
        "TUI_06J": "others",
        "TUI_06E": "parents",
        "TUI_06D": "otherInFAMs",
    },
    2022: {
        "TUI_06A": "Alone",
        "TUI_06B": "Spouse",
        "TUI_06C": "Children",
        "TUI_06H": "friends",
        "TUI_06G": "otherHHs",
        "TUI_06J": "others",
        "TUI_06E": "parents",
        "TUI_06D": "otherInFAMs",
    },
}


def or_merge_copresence(
    df: pd.DataFrame, target_col: str, source_cols: list[str]
) -> pd.DataFrame:
    """OR-merge binary (1/2/NaN) source columns into target_col.

    Rules:
        result = 1   if any source == 1                  (person was present)
        result = 2   if no source == 1, at least one == 2 (person was absent)
        result = NaN if all sources are NaN               (not measured)

    Operates in-place on target_col; drops no columns itself.

    Args:
        df:          Episode DataFrame for one cycle.
        target_col:  Name of the unified column to write the result into.
        source_cols: List of column names to OR-merge (may include target_col itself).

    Returns:
        df with target_col updated.
    """
    available = [c for c in source_cols if c in df.columns]
    if not available:
        return df
    any_present = (df[available] == 1).any(axis=1)
    any_absent  = (df[available] == 2).any(axis=1)
    result = pd.Series(pd.NA, index=df.index, dtype="Int8")
    result[any_present]            = 1
    result[~any_present & any_absent] = 2
    df[target_col] = result
    return df


def harmonize_copresence(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Rename, OR-merge, and clean all co-presence columns.

    Steps:
        A. Standardize missing codes (7/8/9 → NaN) on ALL raw co-presence
           columns (both mapped and unmapped) before any renaming.
        B. Rename primary columns to unified names via COPRESENCE_MAP.
        C. OR-merge unmapped columns into existing unified targets.
        D. Add `colleagues` column (TUI_06I for 2015/2022; NaN for 2005/2010).
        E. Drop all residual raw co-presence columns.

    Produces 9 unified columns:
        Alone, Spouse, Children, parents, otherInFAMs, otherHHs,
        friends, others, colleagues
    """
    rename_map = COPRESENCE_MAP.get(cycle, {})

    # Step A: Standardize missing codes on raw columns before rename
    extra_raw = (
        ["NHSDCL15", "NHSDC15P", "NHSDPAR"] if cycle in (2005, 2010)
        else ["TUI_06F", "TUI_06I"]
    )
    for col in list(rename_map.keys()) + extra_raw:
        if col in df.columns:
            df[col] = df[col].replace({7: pd.NA, 8: pd.NA, 9: pd.NA})

    # Step B: Rename primary columns to unified names
    df = df.rename(columns=rename_map)

    # Step C: OR-merge unmapped columns into unified targets
    if cycle in (2005, 2010):
        # NHSDCL15: children outside HH <15 → merge into Children
        df = or_merge_copresence(df, "Children",    ["Children",    "NHSDCL15"])
        # NHSDPAR:  parents outside HH   → merge into parents
        df = or_merge_copresence(df, "parents",     ["parents",     "NHSDPAR"])
        # NHSDC15P: children outside HH ≥15 → merge into otherInFAMs
        df = or_merge_copresence(df, "otherInFAMs", ["otherInFAMs", "NHSDC15P"])
        # No equivalent for colleagues in 2005/2010
        df["colleagues"] = pd.NA

    else:  # 2015, 2022
        # TUI_06F: other HH adults → merge into otherInFAMs (alongside TUI_06D→otherInFAMs)
        df = or_merge_copresence(df, "otherInFAMs", ["otherInFAMs", "TUI_06F"])
        # TUI_06I: colleagues/classmates → new unified column
        df["colleagues"] = df["TUI_06I"].copy() if "TUI_06I" in df.columns else pd.NA

    # Step D: Drop residual raw co-presence columns
    raw_to_drop = [
        c for c in df.columns
        if c in {"NHSDCL15", "NHSDC15P", "NHSDPAR", "TUI_06F", "TUI_06I"}
    ]
    df = df.drop(columns=raw_to_drop, errors="ignore")

    return df


# --- PHASE F: METADATA & DIARY QA ---


def assign_metadata(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    df["CYCLE_YEAR"] = cycle
    df["SURVYEAR"] = cycle
    if cycle in (2005, 2010):
        df["SURVMNTH"] = pd.NA
    df["COLLECT_MODE"] = 1 if cycle == 2022 else 0
    df["TUI_10_AVAIL"] = 1 if cycle in (2015, 2022) else 0
    df["BS_TYPE"] = "MEAN_BS" if cycle in (2005, 2010) else "STANDARD_BS"
    return df


def convert_hhmm_to_minutes(hhmm_series: pd.Series) -> pd.Series:
    """Convert HHMM integer (e.g., 400 to 240)."""
    hh = hhmm_series // 100
    mm = hhmm_series % 100
    return hh * 60 + mm


def check_diary_closure(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    if "start" not in df.columns or "end" not in df.columns:
        df["DIARY_VALID"] = 0
        return df

    start_raw = pd.to_numeric(df["start"], errors="coerce")
    end_raw = pd.to_numeric(df["end"], errors="coerce")

    if cycle in (2005, 2010):
        # The diary tracks roughly 4:00 AM to 4:00 AM.
        # But respondents might report sleeping from e.g. 23:30 to 07:30 (next day wakes).
        # We must cap the final episode's end time at 4:00 (400) to measure the strict 24h block.
        last_mask = ~df.duplicated(subset=["occID"], keep="last")
        # If the last episode extends past 4AM on the NEXT day (end_raw > 400 but started the night before)
        # Note: In GSS, 4:00 to 23:59 are day 1, 0:00 to 3:59 are day 2.
        # So "400" here means 4:00 AM. Anything from 4:01 to 23:59 represents the previous/current day.
        # Usually, wake-up times like 600 or 730 on the last row mean they slept past the 4AM barrier.
        cap_mask = (
            last_mask & (end_raw > 400) & (end_raw < 1200)
        )  # Only cap morning wake-ups
        end_raw = end_raw.copy()
        end_raw.loc[cap_mask] = 400

        # Also cap ANY very first episode starting before 4:00 AM
        first_mask = ~df.duplicated(subset=["occID"], keep="first")
        cap_start_mask = first_mask & (start_raw < 400)
        start_raw = start_raw.copy()
        start_raw.loc[cap_start_mask] = 400

    start_min = convert_hhmm_to_minutes(start_raw)
    end_min = convert_hhmm_to_minutes(end_raw)

    dur = end_min - start_min
    dur[dur < 0] += 1440  # Wrap midnight

    df["duration"] = dur

    totals = df.groupby("occID")["duration"].sum()
    valid_ids = totals[totals == 1440].index

    df["DIARY_VALID"] = df["occID"].isin(valid_ids).astype(int)
    return df


# --- PHASE G: ORCHESTRATOR ---


def harmonize_main(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    df = recode_sex(df, cycle)
    df = recode_marsth(df, cycle)
    df = recode_agegrp(df, cycle)
    df = recode_lftag(df, cycle)
    df = recode_pr(df, cycle)
    df = recode_cma(df, cycle)
    df = recode_hhsize(df, cycle)
    df = recode_hrswrk(df, cycle)
    df = recode_kol(df, cycle)
    # df = derive_mode(df, cycle)
    df = recode_totinc(df, cycle)
    df = recode_cow(df, cycle)

    for col, sentinels in SENTINEL_MAP.items():
        if col in df.columns:
            df[col] = df[col].replace(list(sentinels), pd.NA)

    df = assign_metadata(df, cycle)
    return df


def harmonize_episode(
    df: pd.DataFrame, cycle: int, act_crosswalk: dict, pre_crosswalk: dict
) -> pd.DataFrame:
    df = apply_activity_crosswalk(df, cycle, act_crosswalk)
    validate_activity_crosswalk(df, cycle, raw_col="ACTCODE" if cycle in (2005, 2010) else "TUI_01")
    df = apply_presence_crosswalk(df, cycle, pre_crosswalk)
    df = harmonize_copresence(df, cycle)
    df = check_diary_closure(df, cycle)

    df["CYCLE_YEAR"] = cycle

    # RF-B1: Drop raw source columns — replaced by derived occACT / occPRE
    df = df.drop(columns=["ACTCODE", "TUI_01", "PLACE", "LOCATION"], errors="ignore")

    # RF-B2: Drop TOTEPISO — 2015-only; derivable from max(EPINO) per occID if needed
    df = df.drop(columns=["TOTEPISO"], errors="ignore")

    # RF-B3: Standardize TUI_07 (tech use) — NaN for 2005/2010 where absent
    if "TUI_07" not in df.columns:
        df["TUI_07"] = pd.NA

    # RF-B4: Unify well-being column → wellbeing (TUI_10 for 2015, TUI_15 for 2022)
    if cycle == 2015:
        if "TUI_10" in df.columns:
            df = df.rename(columns={"TUI_10": "wellbeing"})
            df["wellbeing"] = df["wellbeing"].replace({96: pd.NA, 97: pd.NA, 98: pd.NA, 99: pd.NA})
        else:
            df["wellbeing"] = pd.NA
    elif cycle == 2022:
        if "TUI_15" in df.columns:
            df = df.rename(columns={"TUI_15": "wellbeing"})
            df["wellbeing"] = df["wellbeing"].replace({9: pd.NA})
        else:
            df["wellbeing"] = pd.NA
    else:  # 2005, 2010
        df["wellbeing"] = pd.NA

    return df


def harmonize_all():
    pathlib.Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

    act_crosswalks = build_activity_crosswalks(ACTIVITY_EXCEL)
    pre_crosswalks = build_presence_crosswalks(PRESENCE_EXCEL)

    for cycle in CYCLES:
        print(f"--- Harmonizing {cycle} ---")
        main_path = pathlib.Path(INPUT_DIR) / f"main_{cycle}.csv"
        epi_path = pathlib.Path(INPUT_DIR) / f"episode_{cycle}.csv"

        main_df = pd.read_csv(main_path, low_memory=False)
        epi_df = pd.read_csv(epi_path, low_memory=False)

        h_main = harmonize_main(main_df, cycle)
        h_epi = harmonize_episode(
            epi_df, cycle, act_crosswalks.get(cycle, {}), pre_crosswalks.get(cycle, {})
        )

        out_main = pathlib.Path(OUTPUT_DIR) / f"main_{cycle}.csv"
        out_epi = pathlib.Path(OUTPUT_DIR) / f"episode_{cycle}.csv"

        h_main.to_csv(out_main, index=False)
        h_epi.to_csv(out_epi, index=False)
        print(f"[{cycle}] Main: {h_main.shape}, Epi: {h_epi.shape}")


if __name__ == "__main__":
    harmonize_all()

    import subprocess
    print("\nRunning Step 2 Validation...")
    subprocess.run(["python", "02_harmonizeGSS_val.py"], check=True)
