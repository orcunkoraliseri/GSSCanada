# -*- coding: utf-8 -*-
"""
3rdJ_02_harmonizeGSS_2split.py

Step 2 of the Leg-2 (Residential + Office / AT_WORK) Occupancy Modeling Pipeline:
Data Harmonization.

Ported from:  2J_docs_occ_nTemp/02_harmonizeGSS.py  (Leg-1, verbatim)
Leg-2 deltas (ONLY changes to the Leg-1 version):
  B. AT_WORK flag  — added to apply_presence_crosswalk immediately after AT_HOME:
         df["AT_WORK"] = (df["occPRE"] == 2).astype(int)
     occPRE == 2 = "workplace" for 2005/2010 (PLACE=02); "at work or school"
     for 2015 (LOCATION=301) and 2022 (LOCATION=3301).
     Work-vs-school gating is DEFERRED to the archetype step (Step 5) — the raw
     2015/2022 codes conflate work and school at the location level; they are
     disentangled via the employment-gating variables (MRW_D40B, LFTAG, NOC/NAICS)
     collected at Step 1.  See pipeline doc §2A/§2B for the WFH coding wrinkle:
     in 2022, paid-work-at-home workers are coded LOCATION=3300 (home), so the raw
     AT_WORK rate looks lower (~6%) than the true employed-away rate; this is
     intentionally correct for office-zone BEM (WFH workers are physically absent).

  C. NOCS unification across all four cycles —
     Step 1 renamed 2015/2022 raw columns to NOCS.  Step 2 extends this:
       2005:  SOC91C10        → NOCS
       2010:  NOCS2006_C10    → NOCS
       2015:  NOCS            (already renamed at Step 1)
       2022:  NOCS            (already renamed at Step 1)

  D. NAICS unification across all four cycles —
     Rename per-cycle raw columns to a single NAICS column:
       2005:  NAICS2002_C16   → NAICS
       2010:  NAICS2007_C16   → NAICS
       2015:  NAIC12CY        → NAICS
       2022:  NAIC22CY        → NAICS
     Codes >= 90 (valid-skip / not-applicable sentinels) are set to pd.NA;
     real aggregated NAICS codes are <= ~22 across all cycles.

  E. TELEWORK harmonized binary flag (Int8, {0, 1, NaN}) —
     Instruments differ across cycles (verified at Step 1):
       2005:  no telework variable → all NaN
       2010:  MAR_Q190  usual arrangement Y/N  (1=yes, 2=no; universe {1,2})
       2015:  WTI_130   diary-day measure  (any 1..9 → 1; 96=valid-skip → 0;
              else NaN).  NOTE: 2015 is a DIARY-DAY measure, NOT comparable to
              the 2010/2022 usual Y/N instruments.
       2022:  TLWK_01A  usual Y/N  (1=yes, 2=no; universe {1,2})
     Output column: TELEWORK on the main file.

Everything else (all recode_* functions, activity crosswalk, co-presence OR-merge,
diary closure, metadata flags) is ported VERBATIM from Leg 1.  Residential logic
is bit-identical — only the five labelled deltas above are new.

Paths use a platform-detection block (Windows → cluster → macOS) identical in
structure to 3rdJ_01_readingGSS_2split.py.

Run locally:
    cd Step2_docs
    py -3 -X utf8 3rdJ_02_harmonizeGSS_2split.py

WINDOWS ENCODING NOTE:
    Run with  py -3 -X utf8 3rdJ_02_harmonizeGSS_2split.py
    to avoid cp1252 crashes on UTF-8 console output.
"""

import os
import platform
import pathlib
import pandas as pd
import openpyxl

CYCLES = [2005, 2010, 2015, 2022]

SENTINEL_MAP = {
    "KOL": {7, 8, 9, 98, 99},
    "TOTINC": {98, 99},
    "HRSWRK": {96, 97, 98, 99},
    "NOCS": {96, 97, 98, 99},
    "MARSTH": {8, 9, 99},
    "WKSWRK": {96, 97, 98, 99},
}

# --- PHASE C: DEMOGRAPHIC RECODES (verbatim from Leg 1) ---


def recode_sex(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Rename only processing; ensure values are {1, 2}."""
    return df


def recode_marsth(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Unified values: {1, 2, 3, 4, 5, 6} + NaN."""
    if cycle in (2005, 2010):
        df["MARSTH"] = df["MARSTH"].replace({8: pd.NA, 9: pd.NA})
    elif cycle == 2022:
        df["MARSTH"] = df["MARSTH"].replace({99: pd.NA})
    return df


def recode_agegrp(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """1-7 unified scheme. No changes needed."""
    return df


def recode_lftag(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Collapse labour-force status to a 3-category scheme comparable across all cycles.
    Target scheme:
        1 = Working at paid job (any hours/type; includes vacation, maternity)
        2 = Going to school (student, no paid employment)
        3 = Not employed (retired, HH work, looking for work, illness, other)

    2005/2010 LFSGSS:
        1 (FT employed)       -> 1
        2 (PT employed)       -> 1
        3 (Student+employed)  -> 1
        4 (Student only)      -> 2
        5 (No employment)     -> 3  (undifferentiated; covers retired/HH work/other)
        8/9 (NS/DK)           -> NaN

    2015 ACT7DAYS:
        1 (Working)           -> 1
        2 (Looking for work)  -> 3
        3 (School)            -> 2
        4 (HH work/caring)    -> 3
        5 (Retired)           -> 3
        6 (Other)             -> 3
        97/98/99 (NS/RF/DK)   -> NaN

    2022 ACT7DAYC:
        1 (Working)           -> 1
        2 (School)            -> 2
        3 (HH work/caring)    -> 3
        4 (Retired)           -> 3
        5 (Other)             -> 3
        9 (NS)                -> NaN
    """
    if cycle in (2005, 2010):
        df["LFTAG"] = df["LFTAG"].replace({8: pd.NA, 9: pd.NA})
        df["LFTAG"] = df["LFTAG"].replace({1: 1, 2: 1, 3: 1, 4: 2, 5: 3})
    elif cycle == 2015:
        df["LFTAG"] = df["LFTAG"].replace({97: pd.NA, 98: pd.NA, 99: pd.NA})
        df["LFTAG"] = df["LFTAG"].replace({1: 1, 2: 3, 3: 2, 4: 3, 5: 3, 6: 3})
    elif cycle == 2022:
        df["LFTAG"] = df["LFTAG"].replace({9: pd.NA})
        df["LFTAG"] = df["LFTAG"].replace({1: 1, 2: 2, 3: 3, 4: 3, 5: 3})
    return df


def recode_pr(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """No remap needed. 2005 stays as REGION (1-5), others as PRV (10-59)."""
    return df


def recode_cma(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """LUC_RST 1-3. No remap needed."""
    return df


def recode_hhsize(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Collapse 2005-2015 code 6 into 5."""
    if cycle in (2005, 2010, 2015):
        df["HHSIZE"] = df["HHSIZE"].replace({6: 5})
    return df


def recode_hrswrk(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Map continuous hours to discrete brackets matching 2022 (8 bins)."""
    df["HRSWRK_RAW"] = df["HRSWRK"]
    if cycle in (2005, 2010, 2015):
        df["HRSWRK"] = pd.to_numeric(df["HRSWRK"], errors="coerce")
        df["HRSWRK"] = df["HRSWRK"].where(df["HRSWRK"] <= 75)
        bins = [-1, 29.99, 34.99, 39.99, 44.99, 49.99, 54.99, 59.99, 200]
        labels = [1, 2, 3, 4, 5, 6, 7, 8]
        mask = df["HRSWRK"].notna()
        df.loc[mask, "HRSWRK"] = pd.cut(
            df.loc[mask, "HRSWRK"], bins=bins, labels=labels, right=True
        ).astype("float64")
    elif cycle == 2022:
        df["HRSWRK"] = pd.to_numeric(df["HRSWRK"], errors="coerce")
    return df


def recode_kol(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Collapse language categories to 3 buckets: 1=English, 2=French, 3=Other/Both/Multiple."""
    if cycle in (2005, 2010):
        mapping = {1: 1, 2: 2, 3: 3, 4: 3, 5: 3, 6: 3, 7: 3}
        df["KOL"] = df["KOL"].map(mapping)
    elif cycle in (2015, 2022):
        mapping = {1: 1, 2: 2, 3: 3, 4: 3}
        df["KOL"] = df["KOL"].map(mapping)
    return df


def derive_mode(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Hierarchical priority mapping."""
    if cycle == 2005:
        df["MODE"] = pd.NA
    elif cycle == 2010:
        def get_priority_2010(row):
            def is_checked(val):
                return val in (1, 1.0, "1", "1.0", "Yes", "yes", "YES")
            if is_checked(row.get("CTW_Q140_C01")):
                return 1
            if is_checked(row.get("CTW_Q140_C02")):
                return 2
            if is_checked(row.get("CTW_Q140_C04")) or is_checked(row.get("CTW_Q140_C05")):
                return 3
            if is_checked(row.get("CTW_Q140_C08")):
                return 4
            if is_checked(row.get("CTW_Q140_C07")):
                return 5
            if (
                is_checked(row.get("CTW_Q140_C06"))
                or is_checked(row.get("CTW_Q140_C09"))
                or is_checked(row.get("CTW_Q140_C03"))
            ):
                return 6
            return pd.NA
        df["MODE"] = df.apply(get_priority_2010, axis=1)
    elif cycle == 2015:
        def get_priority_2015(row):
            def is_checked(val):
                return val in (1, 1.0, "1", "1.0", "Yes", "yes", "YES")
            if is_checked(row.get("CTW_140A")):
                return 1
            if is_checked(row.get("CTW_140B")):
                return 2
            if is_checked(row.get("CTW_140C")) or is_checked(row.get("CTW_140D")):
                return 3
            if is_checked(row.get("CTW_140G")):
                return 4
            if is_checked(row.get("CTW_140F")):
                return 5
            if (
                is_checked(row.get("CTW_140E"))
                or is_checked(row.get("CTW_140H"))
                or is_checked(row.get("CTW_140I"))
            ):
                return 6
            return pd.NA
        df["MODE"] = df.apply(get_priority_2015, axis=1)
    elif cycle == 2022:
        df["MODE"] = df["MODE"].replace({99: pd.NA})
    return df


def recode_attsch(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """ATTSCH -- binary school-attendance flag, harmonized across cycles.

    Output: Int8 column ATTSCH in {0, 1, NaN}. Universe-pad-to-No reconciliation --
    respondents outside the source-column universe are mapped to 0 (not 1),
    so cross-cycle prevalence stays comparable. See IMP doc Section 3 Lever A.
    """
    out = pd.Series(pd.NA, index=df.index, dtype="Int8")
    if cycle == 2005:
        if "EDUSTAT" in df.columns:
            s = pd.to_numeric(df["EDUSTAT"], errors="coerce")
            out[s.isin([1, 2])] = 1
            out[s == 7] = 0
    elif cycle == 2010:
        if "EOR_Q320" in df.columns:
            s = pd.to_numeric(df["EOR_Q320"], errors="coerce")
            out[s == 9995] = 1
            out[s.between(1900, 2010)] = 0
    elif cycle == 2015:
        if "ESC1_01" in df.columns:
            s = pd.to_numeric(df["ESC1_01"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
    elif cycle == 2022:
        if "EDC_10" in df.columns:
            s = pd.to_numeric(df["EDC_10"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
    df["ATTSCH"] = out
    return df


def derive_powst(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """POWST -- binary works-from-home flag, harmonized across cycles.

    Output: Int8 column POWST in {0, 1, NaN}. Non-workers stay NaN (LFTAG already
    disambiguates worker vs non-worker upstream -- no POWST_MISSING flag needed).
    """
    out = pd.Series(pd.NA, index=df.index, dtype="Int8")
    if cycle == 2005:
        if "MAR_Q190" in df.columns and "MAR_Q193" in df.columns:
            q190 = pd.to_numeric(df["MAR_Q190"], errors="coerce")
            q193 = pd.to_numeric(df["MAR_Q193"], errors="coerce")
            valid = q190.isin([1, 2])
            out[valid & (q190 == 1) & (q193 == 5)] = 1
            out[valid & ~((q190 == 1) & (q193 == 5))] = 0
    elif cycle == 2010:
        if "CTW_Q140_C08" in df.columns:
            s = pd.to_numeric(df["CTW_Q140_C08"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
    elif cycle == 2015:
        if "CTW_140H" in df.columns:
            s = pd.to_numeric(df["CTW_140H"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
    elif cycle == 2022:
        if "CTW_140I" in df.columns:
            s = pd.to_numeric(df["CTW_140I"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
    df["POWST"] = out
    return df


def recode_totinc(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Income harmonized to buckets. Discretize CRA values in 2022."""
    df["TOTINC_RAW"] = df["TOTINC"]
    if cycle in (2005, 2010):
        df["TOTINC_SOURCE"] = "SELF"
        mapping = {
            1: 1, 2: 1, 3: 1, 4: 1, 5: 1,
            6: 2, 7: 2,
            8: 3, 9: 3,
            10: 4,
            11: 5, 12: 5,
        }
        df["TOTINC"] = df["TOTINC"].map(mapping)
    elif cycle == 2015:
        df["TOTINC_SOURCE"] = "SELF"
        mapping = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 5, 7: 5}
        df["TOTINC"] = df["TOTINC"].map(mapping)
    elif cycle == 2022:
        df["TOTINC_SOURCE"] = "CRA"
    return df


def recode_cow(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Harmonize Class of Worker to 3-category scheme:
        1 = Employee (paid worker for someone else)
        2 = Self-employed (with or without employees) or unpaid family worker
        NaN = Not applicable / not at paid work / sentinel
    """
    if "COW" not in df.columns:
        return df
    if cycle in (2005, 2010):
        df["COW"] = df["COW"].replace({1: 1, 2: 2, 3: 2, 7: pd.NA, 8: pd.NA, 9: pd.NA})
    elif cycle == 2015:
        df["COW"] = df["COW"].replace({1: 2, 2: 1, 6: pd.NA, 7: pd.NA, 8: pd.NA, 9: pd.NA})
    elif cycle == 2022:
        df["COW"] = df["COW"].replace({1: 1, 2: 2, 3: 2, 6: pd.NA, 9: pd.NA})
    return df


# --- DELTA C: NOCS UNIFICATION ---

def unify_nocs(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Leg-2 delta C: unify NOC occupation column to 'NOCS' for all four cycles.

    2015 and 2022 already carry 'NOCS' (renamed at Step 1 from NOC1110Y / NOCLBR_Y).
    2005 and 2010 keep raw names SOC91C10 and NOCS2006_C10 respectively; rename here.
    After rename, apply NOCS sentinel nullification consistent with SENTINEL_MAP
    (codes 96-99 -> NaN) to the newly renamed column.
    """
    if cycle == 2005:
        if "SOC91C10" in df.columns and "NOCS" not in df.columns:
            df = df.rename(columns={"SOC91C10": "NOCS"})
        elif "SOC91C10" in df.columns:
            # NOCS already exists (should not happen); keep NOCS, drop SOC91C10
            df = df.drop(columns=["SOC91C10"], errors="ignore")
    elif cycle == 2010:
        if "NOCS2006_C10" in df.columns and "NOCS" not in df.columns:
            df = df.rename(columns={"NOCS2006_C10": "NOCS"})
        elif "NOCS2006_C10" in df.columns:
            df = df.drop(columns=["NOCS2006_C10"], errors="ignore")
    # 2015/2022: already named NOCS from Step-1 rename map
    if "NOCS" not in df.columns:
        print(f"  [{cycle}] WARN: NOCS column absent after unification attempt -- skipping NOCS for this cycle")
        df["NOCS"] = pd.NA
    # Nullify sentinel codes (not-applicable / don't-know / refused) on the unified column.
    # SOC91C10 (2005) and NOCS2006_C10 (2010) use 97/98/99 as non-response codes.
    # NOC1110Y (2015/2022) uses 95 as valid-skip; 96/97/98/99 are administrative blanks.
    # Set all of {95, 96, 97, 98, 99} -> NaN; legitimate category codes are 1-10 for all cycles.
    NOCS_SENTINELS = {95, 96, 97, 98, 99}
    if "NOCS" in df.columns:
        df["NOCS"] = df["NOCS"].where(~df["NOCS"].isin(NOCS_SENTINELS), other=pd.NA)
    return df


# --- DELTA D: NAICS UNIFICATION ---

def unify_naics(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Leg-2 delta D: unify NAICS industry column to 'NAICS' for all four cycles.

    Per-cycle raw names:
        2005: NAICS2002_C16 -> NAICS
        2010: NAICS2007_C16 -> NAICS
        2015: NAIC12CY      -> NAICS
        2022: NAIC22CY      -> NAICS

    After rename, codes >= 90 (valid-skip / not-applicable sentinels) are set to
    pd.NA.  Real aggregated NAICS codes in the PUMF files are <= ~24 across all
    cycles (the 16-bucket and 22-bucket schemes both top out below 25).
    """
    raw_col_map = {
        2005: "NAICS2002_C16",
        2010: "NAICS2007_C16",
        2015: "NAIC12CY",
        2022: "NAIC22CY",
    }
    raw_col = raw_col_map.get(cycle)
    if raw_col and raw_col in df.columns:
        df = df.rename(columns={raw_col: "NAICS"})
    elif "NAICS" not in df.columns:
        print(f"  [{cycle}] WARN: NAICS source column '{raw_col}' absent -- NAICS will be NaN for this cycle")
        df["NAICS"] = pd.NA

    if "NAICS" in df.columns:
        naics_num = pd.to_numeric(df["NAICS"], errors="coerce")
        df["NAICS"] = naics_num.where(naics_num < 90, other=pd.NA)

    return df


# --- DELTA E: TELEWORK HARMONIZATION ---

def derive_telework(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Leg-2 delta E: harmonize telework into a binary TELEWORK flag (Int8, {0,1,NaN}).

    Instruments differ across cycles:
        2005: MAR_Q190 -- usual paid-work-at-home arrangement Y/N (SAME instrument
              and coding as 2010; present in the 2005 PUMF/main file).
              1 = yes -> TELEWORK = 1
              2 = no  -> TELEWORK = 0
              7/8/9 (not asked / not stated / don't know) -> NaN
        2010: MAR_Q190 -- usual paid-work-at-home arrangement Y/N
              1 = yes -> TELEWORK = 1
              2 = no  -> TELEWORK = 0
              Other / not asked -> NaN
        2015: WTI_130 -- diary-day telework measure (reason for working at home today).
              1..9 = specific work-at-home reason (all mean telework occurred) -> 1
              96   = valid skip (asked, didn't telework) -> 0
              else -> NaN
              NOTE: 2015 is a DIARY-DAY measure.  It captures telework on the
              specific diary day, NOT a usual/habitual arrangement.  It is NOT
              directly comparable to 2010 MAR_Q190 (usual arrangement) or 2022
              TLWK_01A (whether teleworked last week).  Flag this in the validation
              report -- treat the 2015 rate as a diary-day incidence, not a prevalence.
        2022: TLWK_01A -- usual telework last week Y/N
              1 = yes -> TELEWORK = 1
              2 = no  -> TELEWORK = 0
              Other -> NaN

    Output: TELEWORK column on the MAIN file (Int8 nullable integer).
    """
    out = pd.Series(pd.NA, index=df.index, dtype="Int8")

    if cycle in (2005, 2010):
        # MAR_Q190: 1 = yes (works at home in paid job), 2 = no; universe = paid workers.
        # Same instrument and coding in 2005 and 2010 (2005 was previously assumed
        # absent -- it is present in the 2005 PUMF/main file).
        if "MAR_Q190" in df.columns:
            s = pd.to_numeric(df["MAR_Q190"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
        else:
            print(f"  [{cycle}] WARN: MAR_Q190 absent -- TELEWORK will be NaN for {cycle}")

    elif cycle == 2015:
        # WTI_130: diary-day measure (1..9 = worked at home reason; 96 = valid skip)
        if "WTI_130" in df.columns:
            s = pd.to_numeric(df["WTI_130"], errors="coerce")
            out[s.between(1, 9)] = 1
            out[s == 96] = 0
        else:
            print(f"  [{cycle}] WARN: WTI_130 absent -- TELEWORK will be NaN for 2015")

    elif cycle == 2022:
        # TLWK_01A: 1 = teleworked last week, 2 = no; usual/habitual arrangement
        if "TLWK_01A" in df.columns:
            s = pd.to_numeric(df["TLWK_01A"], errors="coerce")
            out[s == 1] = 1
            out[s == 2] = 0
        else:
            print(f"  [{cycle}] WARN: TLWK_01A absent -- TELEWORK will be NaN for 2022")

    df["TELEWORK"] = out
    return df


# --- DELTA F: WORK SCHEDULE / SHIFT-TYPE HARMONIZATION ---

def derive_work_schedule(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Leg-2 delta F: harmonize usual work schedule / shift type into WORK_SCHEDULE.

    Source columns (identical 1-9 category scheme across all four cycles):
        2005 / 2010: MAR_Q410  (97/98/99 = not asked / not stated / don't know)
        2015 / 2022: WHW_230   (96 = valid skip, 97/98/99 = dk / refusal / na)

    Harmonized WORK_SCHEDULE categories (Int8, NaN for sentinels):
        1 = regular daytime schedule / shift
        2 = regular evening shift
        3 = regular night shift
        4 = rotating shift
        5 = split shift
        6 = compressed work week
        7 = on call or casual
        8 = irregular schedule
        9 = other
    Codes are aligned 1:1 between MAR_Q410 and WHW_230, so the raw value is kept
    when in 1..9 and set to NaN otherwise (covers both 96-99 and 97-99 sentinels).
    Universe = workers; non-workers fall outside 1..9 and become NaN.

    Output: WORK_SCHEDULE column on the MAIN file (Int8 nullable integer).
    The raw source column is dropped after derivation.
    """
    raw_col = "MAR_Q410" if cycle in (2005, 2010) else "WHW_230"
    out = pd.Series(pd.NA, index=df.index, dtype="Int8")
    if raw_col in df.columns:
        s = pd.to_numeric(df[raw_col], errors="coerce")
        valid = s.between(1, 9)
        out[valid] = s[valid].astype("Int8")
        df = df.drop(columns=[raw_col], errors="ignore")
    else:
        print(f"  [{cycle}] WARN: {raw_col} absent -- WORK_SCHEDULE will be NaN for {cycle}")
    df["WORK_SCHEDULE"] = out
    return df


# --- PHASE D: ACTIVITY CODE CROSSWALK (verbatim from Leg 1) ---

ACT_SHEET_MAP = {
    2005: "2005codebook",
    2010: "2010codebook",
    2015: "2015codebook",
    2022: "2022codebook",
}

ACT_LABELS = {
    1: "Work & Related",
    2: "Household Work & Maintenance",
    3: "Caregiving & Help",
    4: "Purchasing Goods & Services",
    5: "Sleep & Naps & Resting",
    6: "Eating & Drinking",
    7: "Personal Care",
    8: "Education",
    9: "Socializing",
    10: "Passive Leisure",
    11: "Active Leisure",
    12: "Community & Volunteer",
    13: "Travel",
    14: "Miscellaneous / Idle",
}


def build_activity_crosswalks(excel_path: str) -> dict:
    """Build {cycle_year: {raw_code: category_num}} from the Excel."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    crosswalks = {}
    for cycle_year, sheet_name in ACT_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None or type(row[0]) == str:
                continue
            category = int(row[0])
            raw_str = str(row[1]).replace(",", ".")
            raw_code = float(raw_str) if cycle_year == 2010 else int(float(raw_str))
            lookup[raw_code] = category
            if cycle_year == 2010:
                lookup[int(raw_code)] = category
        crosswalks[cycle_year] = lookup
    return crosswalks


def apply_activity_crosswalk(df: pd.DataFrame, cycle: int, crosswalk: dict) -> pd.DataFrame:
    """Map raw activity codes to 14-category scheme."""
    raw_col = "ACTCODE" if cycle in (2005, 2010) else "TUI_01"
    if raw_col not in df.columns:
        return df

    df["occACT_raw"] = df[raw_col]

    if cycle == 2010:
        df["occACT"] = (
            pd.to_numeric(df[raw_col], errors="coerce").astype(float).map(crosswalk)
        )
        mask = df["occACT"].isna() & df[raw_col].notna()
        if mask.any():
            fallback_map = (
                pd.to_numeric(df.loc[mask, raw_col], errors="coerce")
                .fillna(-1)
                .astype(int)
                .map(crosswalk)
            )
            df.loc[mask, "occACT"] = fallback_map
    else:
        df["occACT"] = pd.to_numeric(df[raw_col], errors="coerce").map(crosswalk)

    if cycle in (2005, 2010):
        df.loc[df[raw_col] == 995, "occACT"] = 10
    elif cycle == 2015:
        df.loc[df[raw_col] == 95, "occACT"] = 14
    elif cycle == 2022:
        df.loc[df[raw_col] == 9999, "occACT"] = 14

    df["occACT_label"] = df["occACT"].map(ACT_LABELS)
    return df


def validate_activity_crosswalk(df: pd.DataFrame, cycle: int, raw_col: str) -> None:
    """Log unmapped raw activity codes with their frequencies."""
    nan_mask = df["occACT"].isna() & df[raw_col].notna()
    if not nan_mask.any():
        print(f"  [{cycle}] All activity codes mapped.")
        return
    unmapped = df.loc[nan_mask, raw_col].value_counts()
    n_eps = nan_mask.sum()
    n_resp = df.loc[nan_mask, "occID"].nunique()
    print(f"  [{cycle}] WARN {n_eps} unmapped episodes in {n_resp} respondents:")
    for code, count in unmapped.items():
        pct = count / len(df) * 100
        print(f"    {raw_col}={code}: {count} episodes ({pct:.2f}%)")


# --- PHASE E: PRESENCE & CO-PRESENCE (verbatim from Leg 1 + AT_WORK delta) ---

PRES_SHEET_MAP = {
    2005: "2005-2010codebook",
    2010: "2005-2010codebook",
    2015: "2015codebook",
    2022: "2022codebook",
}


def build_presence_crosswalks(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    crosswalks = {}
    for cycle_year, sheet_name in PRES_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        lookup = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None or type(row[0]) == str:
                continue
            unified = int(row[0])
            raw_code = int(float(row[1]))
            lookup[raw_code] = unified
        crosswalks[cycle_year] = lookup
    return crosswalks


def apply_presence_crosswalk(df: pd.DataFrame, cycle: int, crosswalk: dict) -> pd.DataFrame:
    """Map raw location codes to unified occPRE scheme and derive AT_HOME + AT_WORK.

    Leg-2 delta B: AT_WORK = (occPRE == 2).astype(int)
    occPRE == 2 maps to:
        2005/2010: PLACE=02  "Work place"          (workplace-only code)
        2015:      LOCATION=301  "At work or school"  (work+school conflated)
        2022:      LOCATION=3301 "At work or school"  (work+school conflated)

    The work-vs-school disambiguation for 2015/2022 is DEFERRED to Step 5
    (archetype linkage), where employment-gating variables (MRW_D40B, LFTAG,
    NOC/NAICS) are available to isolate true-workplace episodes.

    WFH coding (2022): paid-work-at-home is coded LOCATION=3300 (home), NOT 3301,
    so AT_WORK excludes WFH workers by design.  This is the correct behaviour for
    office-zone BEM: a WFH worker is physically absent from the office building.
    """
    raw_col = "PLACE" if cycle in (2005, 2010) else "LOCATION"
    if raw_col not in df.columns:
        return df
    df["occPRE_raw"] = df[raw_col]
    df["occPRE"] = pd.to_numeric(df[raw_col], errors="coerce").map(crosswalk)
    df["AT_HOME"] = (df["occPRE"] == 1).astype(int)
    # Leg-2 delta B: AT_WORK channel
    df["AT_WORK"] = (df["occPRE"] == 2).astype(int)
    return df


COPRESENCE_MAP = {
    2005: {
        "ALONE": "Alone",
        "SPOUSE": "Spouse",
        "CHILDHSD": "Children",
        "FRIENDS": "friends",
        "OTHFAM": "otherHHs",
        "OTHERS": "others",
        "PARHSD": "parents",
        "MEMBHSD": "otherInFAMs",
    },
    2010: {
        "ALONE": "Alone",
        "SPOUSE": "Spouse",
        "CHILDHSD": "Children",
        "FRIENDS": "friends",
        "OTHFAM": "otherHHs",
        "OTHERS": "others",
        "PARHSD": "parents",
        "MEMBHSD": "otherInFAMs",
    },
    2015: {
        "TUI_06A": "Alone",
        "TUI_06B": "Spouse",
        "TUI_06C": "Children",
        "TUI_06H": "friends",
        "TUI_06G": "otherHHs",
        "TUI_06J": "others",
        "TUI_06E": "parents",
        "TUI_06D": "otherInFAMs",
    },
    2022: {
        "TUI_06A": "Alone",
        "TUI_06B": "Spouse",
        "TUI_06C": "Children",
        "TUI_06H": "friends",
        "TUI_06G": "otherHHs",
        "TUI_06J": "others",
        "TUI_06E": "parents",
        "TUI_06D": "otherInFAMs",
    },
}


def or_merge_copresence(df: pd.DataFrame, target_col: str, source_cols: list) -> pd.DataFrame:
    """OR-merge binary (1/2/NaN) source columns into target_col."""
    available = [c for c in source_cols if c in df.columns]
    if not available:
        return df
    any_present = (df[available] == 1).any(axis=1)
    any_absent = (df[available] == 2).any(axis=1)
    result = pd.Series(pd.NA, index=df.index, dtype="Int8")
    result[any_present] = 1
    result[~any_present & any_absent] = 2
    df[target_col] = result
    return df


def harmonize_copresence(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Rename, OR-merge, and clean all co-presence columns."""
    rename_map = COPRESENCE_MAP.get(cycle, {})

    extra_raw = (
        ["NHSDCL15", "NHSDC15P", "NHSDPAR"] if cycle in (2005, 2010)
        else ["TUI_06F", "TUI_06I"]
    )
    for col in list(rename_map.keys()) + extra_raw:
        if col in df.columns:
            df[col] = df[col].replace({7: pd.NA, 8: pd.NA, 9: pd.NA})

    df = df.rename(columns=rename_map)

    if cycle in (2005, 2010):
        df = or_merge_copresence(df, "Children", ["Children", "NHSDCL15"])
        df = or_merge_copresence(df, "parents", ["parents", "NHSDPAR"])
        df = or_merge_copresence(df, "otherInFAMs", ["otherInFAMs", "NHSDC15P"])
        df["colleagues"] = pd.NA
    else:
        df = or_merge_copresence(df, "otherInFAMs", ["otherInFAMs", "TUI_06F"])
        df["colleagues"] = df["TUI_06I"].copy() if "TUI_06I" in df.columns else pd.NA

    raw_to_drop = [
        c for c in df.columns
        if c in {"NHSDCL15", "NHSDC15P", "NHSDPAR", "TUI_06F", "TUI_06I"}
    ]
    df = df.drop(columns=raw_to_drop, errors="ignore")
    return df


# --- PHASE F: METADATA & DIARY QA (verbatim from Leg 1) ---

def assign_metadata(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    df["CYCLE_YEAR"] = cycle
    df["SURVYEAR"] = cycle
    if cycle in (2005, 2010):
        df["SURVMNTH"] = pd.NA
    df["COLLECT_MODE"] = 1 if cycle == 2022 else 0
    df["TUI_10_AVAIL"] = 1 if cycle in (2015, 2022) else 0
    df["BS_TYPE"] = "MEAN_BS" if cycle in (2005, 2010) else "STANDARD_BS"
    return df


def convert_hhmm_to_minutes(hhmm_series: pd.Series) -> pd.Series:
    """Convert HHMM integer (e.g., 400 -> 240)."""
    hh = hhmm_series // 100
    mm = hhmm_series % 100
    return hh * 60 + mm


def check_diary_closure(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    if "start" not in df.columns or "end" not in df.columns:
        df["DIARY_VALID"] = 0
        return df

    start_raw = pd.to_numeric(df["start"], errors="coerce")
    end_raw = pd.to_numeric(df["end"], errors="coerce")

    if cycle in (2005, 2010):
        last_mask = ~df.duplicated(subset=["occID"], keep="last")
        cap_mask = last_mask & (end_raw > 400) & (end_raw < 1200)
        end_raw = end_raw.copy()
        end_raw.loc[cap_mask] = 400

        first_mask = ~df.duplicated(subset=["occID"], keep="first")
        cap_start_mask = first_mask & (start_raw < 400)
        start_raw = start_raw.copy()
        start_raw.loc[cap_start_mask] = 400

    start_min = convert_hhmm_to_minutes(start_raw)
    end_min = convert_hhmm_to_minutes(end_raw)

    dur = end_min - start_min
    dur[dur < 0] += 1440

    df["duration"] = dur

    totals = df.groupby("occID")["duration"].sum()
    valid_ids = totals[totals == 1440].index

    df["DIARY_VALID"] = df["occID"].isin(valid_ids).astype(int)
    return df


# --- PHASE G: ORCHESTRATORS ---

def harmonize_main(df: pd.DataFrame, cycle: int) -> pd.DataFrame:
    """Apply all main-file harmonization steps (residential verbatim + Leg-2 deltas C/D/E)."""
    # Residential recodes (verbatim from Leg 1)
    df = recode_sex(df, cycle)
    df = recode_marsth(df, cycle)
    df = recode_agegrp(df, cycle)
    df = recode_lftag(df, cycle)
    df = recode_pr(df, cycle)
    df = recode_cma(df, cycle)
    df = recode_hhsize(df, cycle)
    df = recode_hrswrk(df, cycle)
    df = recode_kol(df, cycle)
    df = derive_mode(df, cycle)
    df = recode_attsch(df, cycle)
    df = derive_powst(df, cycle)
    df = recode_totinc(df, cycle)
    df = recode_cow(df, cycle)

    for col, sentinels in SENTINEL_MAP.items():
        if col in df.columns:
            df[col] = df[col].replace(list(sentinels), pd.NA)

    df = assign_metadata(df, cycle)

    # Leg-2 deltas C, D, E, F
    df = unify_nocs(df, cycle)
    df = unify_naics(df, cycle)
    df = derive_telework(df, cycle)
    df = derive_work_schedule(df, cycle)

    return df


def harmonize_episode(
    df: pd.DataFrame,
    cycle: int,
    act_crosswalk: dict,
    pre_crosswalk: dict,
) -> pd.DataFrame:
    """Apply all episode-file harmonization steps (residential verbatim + AT_WORK delta B)."""
    df = apply_activity_crosswalk(df, cycle, act_crosswalk)
    validate_activity_crosswalk(
        df, cycle, raw_col="ACTCODE" if cycle in (2005, 2010) else "TUI_01"
    )
    # Leg-2 delta B is inside apply_presence_crosswalk:
    #   AT_WORK = (occPRE == 2).astype(int)
    df = apply_presence_crosswalk(df, cycle, pre_crosswalk)
    df = harmonize_copresence(df, cycle)
    df = check_diary_closure(df, cycle)

    df["CYCLE_YEAR"] = cycle

    # RF-B1: Drop raw source columns
    df = df.drop(columns=["ACTCODE", "TUI_01", "PLACE", "LOCATION"], errors="ignore")

    # RF-B2: Drop TOTEPISO
    df = df.drop(columns=["TOTEPISO"], errors="ignore")

    # RF-B3: Standardize TUI_07
    if "TUI_07" not in df.columns:
        df["TUI_07"] = pd.NA

    # RF-B4: Unify well-being column
    if cycle == 2015:
        if "TUI_10" in df.columns:
            df = df.rename(columns={"TUI_10": "wellbeing"})
            df["wellbeing"] = df["wellbeing"].replace({96: pd.NA, 97: pd.NA, 98: pd.NA, 99: pd.NA})
        else:
            df["wellbeing"] = pd.NA
    elif cycle == 2022:
        if "TUI_15" in df.columns:
            df = df.rename(columns={"TUI_15": "wellbeing"})
            df["wellbeing"] = df["wellbeing"].replace({9: pd.NA})
        else:
            df["wellbeing"] = pd.NA
    else:
        df["wellbeing"] = pd.NA

    return df


def harmonize_all(
    input_dir: str,
    output_dir: str,
    activity_excel: str,
    presence_excel: str,
) -> None:
    """Run the full harmonization pipeline for all four cycles."""
    pathlib.Path(output_dir).mkdir(parents=True, exist_ok=True)

    print("Loading activity crosswalk from Excel...")
    act_crosswalks = build_activity_crosswalks(activity_excel)
    print("Loading presence crosswalk from Excel...")
    pre_crosswalks = build_presence_crosswalks(presence_excel)

    for cycle in CYCLES:
        print(f"\n--- Harmonizing {cycle} ---")
        main_path = pathlib.Path(input_dir) / f"main_{cycle}.csv"
        epi_path = pathlib.Path(input_dir) / f"episode_{cycle}.csv"

        main_df = pd.read_csv(main_path, low_memory=False)
        epi_df = pd.read_csv(epi_path, low_memory=False)

        h_main = harmonize_main(main_df, cycle)
        h_epi = harmonize_episode(
            epi_df, cycle,
            act_crosswalks.get(cycle, {}),
            pre_crosswalks.get(cycle, {}),
        )

        out_main = pathlib.Path(output_dir) / f"main_{cycle}.csv"
        out_epi = pathlib.Path(output_dir) / f"episode_{cycle}.csv"

        h_main.to_csv(out_main, index=False)
        h_epi.to_csv(out_epi, index=False)
        print(f"  [{cycle}] Main: {h_main.shape}  |  Episode: {h_epi.shape}")

        # Spot-check AT_WORK rate (episode-weighted)
        if "AT_WORK" in h_epi.columns:
            rate = h_epi["AT_WORK"].mean() * 100
            print(f"  [{cycle}] AT_WORK rate (episode-level, unweighted): {rate:.2f}%")
        if "TELEWORK" in h_main.columns:
            tw_rate = h_main["TELEWORK"].dropna().mean() * 100
            tw_n = h_main["TELEWORK"].notna().sum()
            print(f"  [{cycle}] TELEWORK rate (among non-NaN, n={tw_n:,}): {tw_rate:.2f}%")
        if "NOCS" in h_main.columns:
            nocs_cov = h_main["NOCS"].notna().mean() * 100
            print(f"  [{cycle}] NOCS coverage: {nocs_cov:.1f}%")
        if "NAICS" in h_main.columns:
            naics_cov = h_main["NAICS"].notna().mean() * 100
            print(f"  [{cycle}] NAICS coverage: {naics_cov:.1f}%")


# --- ENTRY POINT ---

if __name__ == "__main__":
    # Platform-detection block (same ordering as 3rdJ_01_readingGSS_2split.py)
    _WIN_BASE = r"C:\Users\o_iseri\Desktop\GSSCanada\GSSCanada-main"
    _SPEED_BASE = "/speed-scratch/o_iseri/GSSCanada/GSSCanada-main"
    _MAC_BASE = "/Users/orcunkoraliseri/Desktop/Postdoc/occModeling"

    if platform.system() == "Windows":
        _BASE = _WIN_BASE
        _SEP = "\\"
    elif os.path.isdir("/speed-scratch/o_iseri"):
        _BASE = _SPEED_BASE
        _SEP = "/"
    else:
        _BASE = _MAC_BASE
        _SEP = "/"

    def _p(*parts):
        return os.path.join(_BASE, *parts)

    INPUT_DIR = _p(
        "3J_docs_occ_nTemp", "Leg2_2-split", "Step1_docs", "outputs_step1"
    )
    OUTPUT_DIR = _p(
        "3J_docs_occ_nTemp", "Leg2_2-split", "Step2_docs", "outputs_step2"
    )
    ACTIVITY_EXCEL = _p(
        "2J_docs_occ_nTemp",
        "references_activityCodes",
        "Data Harmonization_activityCategories - execution.xlsx",
    )
    PRESENCE_EXCEL = _p(
        "2J_docs_occ_nTemp",
        "references_Pre_coPre_Codes",
        "Data Harmonization_presenceCategories - execution.xlsx",
    )

    print("=== 3rdJ Leg-2 Step 2: Data Harmonization ===")
    print(f"  Input:          {INPUT_DIR}")
    print(f"  Output:         {OUTPUT_DIR}")
    print(f"  Activity Excel: {ACTIVITY_EXCEL}")
    print(f"  Presence Excel: {PRESENCE_EXCEL}")
    print()

    harmonize_all(INPUT_DIR, OUTPUT_DIR, ACTIVITY_EXCEL, PRESENCE_EXCEL)

    print("\n=== Harmonization complete. Running Step 2 Validator... ===")
    import subprocess
    import sys
    script_dir = os.path.dirname(os.path.abspath(__file__))
    val_script = os.path.join(script_dir, "3rdJ_02_harmonizeGSS_2split_val.py")
    result = subprocess.run(
        [sys.executable, "-X", "utf8", val_script],
        check=False,
    )
    if result.returncode != 0:
        print(f"WARN: validator exited with code {result.returncode} -- inspect HTML report")
    else:
        print("Validator completed successfully.")
